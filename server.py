"""
Equity Dashboard backend — pulls REAL data for BSE / NSE listed companies.

Source: Yahoo Finance via `yfinance` (mirrors NSE `.NS` and BSE `.BO` tickers).
Yahoo carries live price + full cash-flow / balance-sheet / income statements,
which is what the dashboard's panels need. (NSE/BSE and Screener.in block
cross-origin browser calls and/or need a login, so a small backend like this
is the practical way to serve real data to a browser page.)

Run:
    python server.py
Then open http://127.0.0.1:5000  in your browser.
"""
import math
import os
import re
import datetime as dt
from flask import Flask, jsonify, send_from_directory, request
from flask_cors import CORS
import yfinance as yf
import nse  # NSE-first client (falls back gracefully when NSE is blocked)
import framework  # reads the investor's audited portfolio Excel
import sectors  # sector-specific analytical lenses
import screener_web  # auto-fetch quarterly financials from Screener.in (public)

app = Flask(__name__, static_folder=".")
CORS(app)

CR = 1e7  # 1 crore = 10,000,000  -> convert absolute INR to ₹ Cr

# ---- a curated NSE symbol list for the search box (type any ticker too) ----
COMPANIES = [
    ("Reliance Industries", "RELIANCE"), ("Tata Consultancy Services", "TCS"),
    ("HDFC Bank", "HDFCBANK"), ("ICICI Bank", "ICICIBANK"), ("Infosys", "INFY"),
    ("Bharti Airtel", "BHARTIARTL"), ("State Bank of India", "SBIN"),
    ("Life Insurance Corp", "LICI"), ("ITC", "ITC"), ("Hindustan Unilever", "HINDUNILVR"),
    ("Larsen & Toubro", "LT"), ("Bajaj Finance", "BAJFINANCE"), ("HCL Technologies", "HCLTECH"),
    ("Maruti Suzuki", "MARUTI"), ("Sun Pharma", "SUNPHARMA"), ("Kotak Mahindra Bank", "KOTAKBANK"),
    ("Axis Bank", "AXISBANK"), ("NTPC", "NTPC"), ("Mahindra & Mahindra", "M&M"),
    ("Titan Company", "TITAN"), ("UltraTech Cement", "ULTRACEMCO"), ("Asian Paints", "ASIANPAINT"),
    ("Adani Enterprises", "ADANIENT"), ("Adani Ports", "ADANIPORTS"), ("Wipro", "WIPRO"),
    ("Power Grid Corp", "POWERGRID"), ("Nestle India", "NESTLEIND"), ("Coal India", "COALINDIA"),
    ("Bajaj Finserv", "BAJAJFINSV"), ("JSW Steel", "JSWSTEEL"), ("Tata Steel", "TATASTEEL"),
    ("Tata Motors", "TATAMOTORS"), ("Oil & Natural Gas Corp", "ONGC"), ("Hindalco", "HINDALCO"),
    ("Grasim Industries", "GRASIM"), ("Tech Mahindra", "TECHM"), ("IndusInd Bank", "INDUSINDBK"),
    ("Britannia Industries", "BRITANNIA"), ("Cipla", "CIPLA"), ("Dr Reddy's Labs", "DRREDDY"),
    ("Eicher Motors", "EICHERMOT"), ("Apollo Hospitals", "APOLLOHOSP"), ("Bajaj Auto", "BAJAJ-AUTO"),
    ("Divi's Labs", "DIVISLAB"), ("Hero MotoCorp", "HEROMOTOCO"), ("SBI Life Insurance", "SBILIFE"),
    ("HDFC Life Insurance", "HDFCLIFE"), ("BPCL", "BPCL"), ("Tata Consumer", "TATACONSUM"),
    ("Shriram Finance", "SHRIRAMFIN"), ("Pidilite Industries", "PIDILITIND"),
    ("DMart (Avenue Supermarts)", "DMART"), ("Adani Green Energy", "ADANIGREEN"),
    ("Zomato", "ZOMATO"), ("Vedanta", "VEDL"), ("DLF", "DLF"), ("Godrej Consumer", "GODREJCP"),
    ("Havells India", "HAVELLS"), ("Siemens", "SIEMENS"), ("Bosch", "BOSCHLTD"),
    ("Varun Beverages", "VBL"), ("TVS Motor", "TVSMOTOR"), ("ABB India", "ABB"),
    ("Ambuja Cements", "AMBUJACEM"), ("Berger Paints", "BERGEPAINT"), ("Marico", "MARICO"),
    ("Dabur India", "DABUR"), ("United Spirits", "MCDOWELL-N"), ("Colgate Palmolive", "COLPAL"),
    ("Trent", "TRENT"), ("Jindal Steel", "JINDALSTEL"), ("Bank of Baroda", "BANKBARODA"),
    ("Punjab National Bank", "PNB"), ("Canara Bank", "CANBK"), ("GAIL India", "GAIL"),
    ("Indian Oil Corp", "IOC"), ("Interglobe Aviation (IndiGo)", "INDIGO"),
    ("Info Edge", "NAUKRI"), ("LTIMindtree", "LTIM"), ("Persistent Systems", "PERSISTENT"),
    ("Polycab India", "POLYCAB"), ("SRF", "SRF"), ("Page Industries", "PAGEIND"),
    ("Mphasis", "MPHASIS"), ("Muthoot Finance", "MUTHOOTFIN"), ("Bajaj Holdings", "BAJAJHLDNG"),
    ("Jio Financial Services", "JIOFIN"), ("Union Bank of India", "UNIONBANK"),
    ("IRCTC", "IRCTC"), ("Tata Power", "TATAPOWER"), ("PI Industries", "PIIND"),
    ("Torrent Pharma", "TORNTPHARM"), ("Max Healthcare", "MAXHEALTH"),
]
NAME_BY_SYM = {s: n for n, s in COMPANIES}


def safe_row(df, *names):
    """Return the first matching row (a pandas Series across periods) or None."""
    if df is None or df.empty:
        return None
    for nm in names:
        if nm in df.index:
            return df.loc[nm]
    return None


def val(series, i=0):
    """Latest (i=0) numeric value from a Series, or None."""
    if series is None:
        return None
    try:
        v = series.iloc[i]
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return None
        return float(v)
    except Exception:
        return None


def cr(x):
    return None if x is None else round(x / CR)


def resolve_ticker(sym: str):
    """Try NSE (.NS) first, then BSE (.BO). Accept a full ticker as given."""
    sym = sym.strip().upper()
    candidates = []
    if sym.endswith((".NS", ".BO")):
        candidates = [sym]
    else:
        candidates = [sym + ".NS", sym + ".BO"]
    for c in candidates:
        t = yf.Ticker(c)
        try:
            h = t.history(period="7d")
            if h is not None and not h.empty:
                return t, c, h
        except Exception:
            continue
    return None, None, None


def build_company(sym: str):
    t, resolved, hist = resolve_ticker(sym)
    if t is None:
        return None

    info = {}
    try:
        info = t.info or {}
    except Exception:
        info = {}

    base_sym = resolved.replace(".NS", "").replace(".BO", "")
    exch = "NSE" if resolved.endswith(".NS") else "BSE"

    # ---- price: NSE first (authoritative), else Yahoo history ----
    # Yahoo can return a NaN for the latest (still-forming) bar, so use the last
    # VALID close and fall back to fast_info/info if the whole series is empty.
    closes = hist["Close"].dropna()
    if len(closes):
        price = round(float(closes.iloc[-1]), 2)
        prev = round(float(closes.iloc[-2]), 2) if len(closes) > 1 else price
    else:
        price = prev = None
    if price is None or (isinstance(price, float) and math.isnan(price)):
        for k in ("last_price", "regularMarketPrice", "currentPrice", "previousClose"):
            cand = info.get(k) if isinstance(info, dict) else None
            try:
                if cand is not None and not math.isnan(float(cand)):
                    price = round(float(cand), 2)
                    break
            except (TypeError, ValueError):
                continue
        if prev is None:
            prev = price
    price_source = "Yahoo Finance"
    nse_name = nse_sector = None
    nse_hi = nse_lo = None
    if resolved.endswith(".NS"):
        npb = nse.price_bundle(base_sym)
        if npb and npb.get("price"):
            price = round(float(npb["price"]), 2)
            prev = round(float(npb["prevClose"]), 2) if npb.get("prevClose") else prev
            nse_hi, nse_lo = npb.get("high52"), npb.get("low52")
            nse_name, nse_sector = npb.get("name"), npb.get("sector")
            price_source = "NSE (live)"

    cf = t.cashflow
    bs = t.balance_sheet
    inc = t.financials

    # ---------- cash-flow lines ----------
    ocf_s = safe_row(cf, "Operating Cash Flow", "Cash Flow From Continuing Operating Activities")
    capex_s = safe_row(cf, "Capital Expenditure")
    fcf_s = safe_row(cf, "Free Cash Flow")
    endcash_s = safe_row(cf, "End Cash Position")
    issdebt_s = safe_row(cf, "Issuance Of Debt", "Long Term Debt Issuance")
    repdebt_s = safe_row(cf, "Repayment Of Debt", "Long Term Debt Payments")
    intpaid_s = safe_row(cf, "Interest Paid Cff")
    div_s = safe_row(cf, "Cash Dividends Paid", "Common Stock Dividend Paid")

    ocf = val(ocf_s)
    capex = abs(val(capex_s)) if val(capex_s) is not None else None
    fcf = val(fcf_s)
    if fcf is None and ocf is not None and capex is not None:
        fcf = ocf - capex
    endcash = val(endcash_s)

    kpis = {
        "opCashFlow": cr(ocf), "capex": cr(capex),
        "cashBal": cr(endcash), "fcf": cr(fcf),
    }

    # ---------- sources / uses ----------
    sources = []
    if ocf and ocf > 0:
        sources.append({"k": "Operating cash flow", "v": cr(ocf), "c": "green"})
    if val(issdebt_s):
        sources.append({"k": "New borrowings", "v": cr(abs(val(issdebt_s))), "c": "amber"})
    uses = []
    if capex:
        uses.append({"k": "Capex", "v": cr(capex), "c": "violet"})
    if val(repdebt_s):
        uses.append({"k": "Debt repayment", "v": cr(abs(val(repdebt_s))), "c": "amber"})
    if val(intpaid_s):
        uses.append({"k": "Interest paid", "v": cr(abs(val(intpaid_s))), "c": "red"})
    if val(div_s):
        uses.append({"k": "Dividends", "v": cr(abs(val(div_s))), "c": "blue"})

    # ---------- multi-year capex vs ocf, cash & fcf ----------
    capex_rows, fcf_rows = [], []
    if cf is not None and not cf.empty:
        cols = list(cf.columns)[:5][::-1]  # oldest -> newest
        for col in cols:
            y = "FY" + str(col.year % 100).zfill(2)
            o = cf.loc["Operating Cash Flow", col] if "Operating Cash Flow" in cf.index else None
            cx = cf.loc["Capital Expenditure", col] if "Capital Expenditure" in cf.index else None
            f = cf.loc["Free Cash Flow", col] if "Free Cash Flow" in cf.index else None
            ec = cf.loc["End Cash Position", col] if "End Cash Position" in cf.index else None
            if o is not None and cx is not None and not (math.isnan(o) or math.isnan(cx)):
                capex_rows.append({"y": y, "ocf": cr(float(o)), "capex": cr(abs(float(cx)))})
            if f is not None and ec is not None and not (math.isnan(f) or math.isnan(ec)):
                fcf_rows.append({"y": y, "cash": cr(float(ec)), "fcf": cr(float(f))})

    # ---------- working capital ----------
    rev_s = safe_row(inc, "Total Revenue", "Operating Revenue")
    cogs_s = safe_row(inc, "Cost Of Revenue", "Reconciled Cost Of Revenue")
    recv_s = safe_row(bs, "Receivables", "Accounts Receivable", "Gross Accounts Receivable")
    inv_s = safe_row(bs, "Inventory")
    pay_s = safe_row(bs, "Payables", "Accounts Payable", "Payables And Accrued Expenses")

    def days(num, den):
        n, d = val(num), val(den)
        if not n or not d or d == 0:
            return None
        return round(n / d * 365)

    dso = days(recv_s, rev_s)
    dio = days(inv_s, cogs_s)
    dpo = days(pay_s, cogs_s)
    wc = {"dso": dso or 0, "dio": dio or 0, "dpo": dpo or 0}

    # ccc per year (align annual columns)
    ccc_rows = []
    if inc is not None and bs is not None and not inc.empty and not bs.empty:
        common = [c for c in inc.columns if c in bs.columns][:5][::-1]
        for col in common:
            def d2(bsrow, incrow):
                try:
                    n = bsrow[col]; dd = incrow[col]
                    if n is None or dd in (None, 0) or math.isnan(n) or math.isnan(dd):
                        return None
                    return n / dd * 365
                except Exception:
                    return None
            r = rev_s if rev_s is not None else None
            cg = cogs_s if cogs_s is not None else None
            dso_y = d2(recv_s, r) if (recv_s is not None and r is not None) else None
            dio_y = d2(inv_s, cg) if (inv_s is not None and cg is not None) else None
            dpo_y = d2(pay_s, cg) if (pay_s is not None and cg is not None) else None
            if dso_y is not None:
                ccc = round((dso_y or 0) + (dio_y or 0) - (dpo_y or 0))
                ccc_rows.append({"y": "FY" + str(col.year % 100).zfill(2), "d": ccc})

    # ---------- quarters (last 8) ----------
    quarters = []
    qf = t.quarterly_financials
    if qf is not None and not qf.empty:
        def qget(row, col):
            v = qf.loc[row, col] if row in qf.index else None
            return None if (v is None or (isinstance(v, float) and math.isnan(v))) else float(v)
        # Indian fiscal quarters (FY = Apr-Mar): quarter-end month -> (fiscal Q, FY)
        # Jun=Q1, Sep=Q2, Dec=Q3, Mar=Q4; FY is the year the fiscal year ENDS.
        FQ = {6: 1, 9: 2, 12: 3, 3: 4}

        def fiscal(mon, yr):
            fq = FQ.get(mon) or ((mon - 1) // 3 + 1)
            fy = yr + 1 if mon >= 4 else yr
            return fq, fy

        def flabel(o):
            fy, fq = o // 4, o % 4 + 1
            return f"Q{fq} FY{str(fy % 100).zfill(2)}"

        # collect available quarters keyed by a fiscal ordinal (fy*4 + fq-1),
        # so we can emit a CONSECUTIVE series and expose any gaps Yahoo left.
        avail = {}
        for col in qf.columns:
            try:
                mon, yr = col.month, col.year
            except Exception:
                continue
            fq, fy = fiscal(mon, yr)
            o = fy * 4 + (fq - 1)
            r = qget("Total Revenue", col) or qget("Operating Revenue", col)
            n = qget("Net Income", col)
            if r is None or n is None:
                continue
            e = qget("EBITDA", col)
            if e is None:  # reconstruct EBITDA = EBIT + D&A if not reported
                eb, da = qget("EBIT", col), qget("Reconciled Depreciation", col)
                e = (eb + da) if (eb is not None and da is not None) else None
            intr = qget("Interest Expense", col) or qget("Interest Expense Non Operating", col)
            avail[o] = {"rev": cr(r), "np": cr(n),
                        "ebitda": cr(e) if e is not None else None,
                        "interest": cr(abs(intr)) if intr is not None else None}
        if avail:
            omax, omin = max(avail), min(avail)
            start = max(omin, omax - 7)   # up to 8 consecutive quarters, none before real data
            # fill any gaps Yahoo left by auto-fetching Screener.in (public, complete series)
            gaps = [o for o in range(start, omax + 1) if o not in avail]
            if gaps:
                try:
                    sq = screener_web.fetch_quarterly(base_sym)
                except Exception:
                    sq = None
                if sq:
                    smap = {(x["y"], x["m"]): x for x in sq}
                    emap = {1: 6, 2: 9, 3: 12, 4: 3}   # fiscal Q -> quarter-end month
                    for o in gaps:
                        fy, fq = o // 4, o % 4 + 1
                        emonth = emap[fq]
                        eyear = fy - 1 if fq <= 3 else fy
                        x = smap.get((eyear, emonth))
                        if x and x.get("rev") is not None:   # Screener values already in ₹ Cr
                            avail[o] = {"rev": round(x["rev"]),
                                        "np": round(x["np"]) if x.get("np") is not None else None,
                                        "ebitda": round(x["ebitda"]) if x.get("ebitda") is not None else None,
                                        "interest": round(x["interest"]) if x.get("interest") is not None else None,
                                        "src": "Screener"}
            for o in range(start, omax + 1):
                if o in avail:
                    quarters.append({"q": flabel(o), "o": o, **avail[o], "missing": False})
                else:
                    quarters.append({"q": flabel(o), "o": o, "rev": None, "np": None,
                                     "ebitda": None, "interest": None, "missing": True})

    # ---------- ratios ----------
    pe = info.get("trailingPE")
    pb = info.get("priceToBook")
    dy = info.get("dividendYield")
    mcap = info.get("marketCap")
    roe = info.get("returnOnEquity")
    npm = info.get("profitMargins")
    evebitda = info.get("enterpriseToEbitda")

    equity = val(safe_row(bs, "Stockholders Equity", "Common Stock Equity"))
    ebit = val(safe_row(inc, "EBIT"))
    inv_cap = val(safe_row(bs, "Invested Capital"))
    roce = (ebit / inv_cap * 100) if (ebit and inv_cap) else None
    ni = val(safe_row(inc, "Net Income"))
    roe_calc = (ni / equity * 100) if (ni and equity) else (roe * 100 if roe else None)

    ratios = []
    if roe_calc is not None: ratios.append({"k": "ROE", "v": f"{roe_calc:.1f}%"})
    if roce is not None: ratios.append({"k": "ROCE", "v": f"{roce:.1f}%"})
    if npm is not None: ratios.append({"k": "Net margin", "v": f"{npm*100:.1f}%"})
    if pe: ratios.append({"k": "P/E", "v": f"{pe:.1f}"})
    if pb: ratios.append({"k": "P/B", "v": f"{pb:.1f}"})
    if evebitda: ratios.append({"k": "EV/EBITDA", "v": f"{evebitda:.1f}"})

    # ---------- debt ----------
    total_debt = val(safe_row(bs, "Total Debt")) or info.get("totalDebt")
    net_debt = val(safe_row(bs, "Net Debt"))
    if net_debt is None and total_debt is not None and endcash is not None:
        net_debt = total_debt - endcash
    de = (total_debt / equity) if (total_debt is not None and equity) else None
    int_exp = val(safe_row(inc, "Interest Expense"))
    icov = (ebit / int_exp) if (ebit and int_exp) else None
    debt = {
        "grossDebt": cr(total_debt), "netDebt": cr(net_debt),
        "deRatio": round(de, 2) if de is not None else 0,
        "interestCover": round(icov, 1) if icov else 99,
        "rating": "—",
    }

    # Screener-stated financials (deep annual history + balance-sheet/ratio fallbacks),
    # fetched once here and reused by valuation and the forensic module below.
    try:
        sfin_data = screener_web.fetch_financials(base_sym)
    except Exception:
        sfin_data = None

    # ---------- intrinsic value (method chosen to fit the business type) ----------
    shares = val(safe_row(bs, "Ordinary Shares Number", "Share Issued"))
    ebitda_v = val(safe_row(inc, "EBITDA", "Normalized EBITDA")) or info.get("ebitda")
    netppe = val(safe_row(bs, "Net PPE"))
    rev_now = val(rev_s)
    sector_txt = info.get("sector") or ""
    blob = (sector_txt + " " + (info.get("longName") or base_sym)).lower()
    is_financial = (sector_txt == "Financial Services") or any(
        k in blob for k in ["bank", "insurance", " finance", "financial", "nbfc"])
    cap_intensity = (netppe / rev_now) if (netppe and rev_now) else None
    is_heavy = (sector_txt in ("Energy", "Utilities", "Basic Materials", "Real Estate")) \
        or (cap_intensity is not None and cap_intensity > 0.8) \
        or (fcf is not None and ocf and fcf < 0.30 * ocf)

    intrinsic, method = None, ""
    scen = None                 # {bear, base, bull} IV band from the sector-appropriate method
    # growth estimate from revenue CAGR
    g = 0.08
    if rev_s is not None and len(rev_s) >= 3:
        old = val(rev_s, min(3, len(rev_s) - 1)); new = val(rev_s, 0)
        yrs = min(3, len(rev_s) - 1)
        if old and new and old > 0 and yrs > 0:
            g = max(0.03, min(0.15, (new / old) ** (1 / yrs) - 1))
    wacc = 0.11
    COE = 0.12                      # cost of equity for Indian large-caps (~12%)
    bvps = (equity / shares) if (equity and shares) else None
    roe_frac = (roe_calc / 100) if roe_calc is not None else None

    # ---- normalised (through-cycle) earnings — never value a trough or a peak on trailing ----
    # (Framework §7.2: cyclicals/recovery names must be valued on mid-cycle, not spot, earnings.)
    import statistics as _stats
    ni_hist = [v for v in (val(safe_row(inc, "Net Income"), i) for i in range(5)) if v is not None]
    ebitda_hist = [v for v in (val(safe_row(inc, "EBITDA", "Normalized EBITDA"), i) for i in range(5)) if v is not None]
    rev_hist = [v for v in (val(rev_s, i) for i in range(5)) if v is not None] if rev_s is not None else []
    # Use Screener's annual history (cleaner & consolidated), but over the RECENT window only
    # (last ~6 yrs): a full 12-yr median would under-normalise a company that has grown, while a
    # 6-yr window still smooths a one-off trough/peak. Screener P&L is ₹ Cr → convert to raw ₹.
    _WIN = 6
    _sh = sfin_data.get("npHist") if isinstance(sfin_data, dict) else None
    if _sh and len(_sh) >= 4:
        ni_hist = [v * 1e7 for v in _sh[-_WIN:]]
    _oh = sfin_data.get("opHist") if isinstance(sfin_data, dict) else None
    if _oh and len(_oh) >= 4:
        ebitda_hist = [v * 1e7 for v in _oh[-_WIN:]]
    _rh = sfin_data.get("salesHist") if isinstance(sfin_data, dict) else None
    if _rh and len(_rh) >= 4:
        rev_hist = [v * 1e7 for v in _rh[-_WIN:]]
    ni_trail = ni
    ni_norm = _stats.median(ni_hist) if len(ni_hist) >= 3 else ni_trail
    rev_intact = bool(rev_now and rev_hist and rev_now >= 0.85 * max(rev_hist))
    earn_state, base_ni = "trailing", ni_trail
    if ni_norm and ni_norm > 0 and ni_trail is not None:
        if ni_trail < 0.6 * ni_norm and rev_intact:
            base_ni, earn_state = ni_norm, "mid-cycle · trough year normalised"
        elif ni_trail > 1.6 * ni_norm:
            base_ni, earn_state = ni_norm, "mid-cycle · peak year normalised"
    fwd_eps = info.get("forwardEps")

    # sector-appropriate P/E band (low, base, high) — the framework's "no single hammer" rule
    _prof = sectors.profile(sector_txt, info.get("longName") or base_sym)
    _bucket = _prof.get("bucket", "")
    _PE_BANDS = {
        "IT Services": (18, 24, 30), "Pharma & Healthcare": (20, 26, 32),
        "FMCG / Consumer Staples": (32, 42, 52), "Consumer Discretionary / Retail": (30, 42, 58),
        "Capital Goods / Engineering / Infra / Defence": (18, 24, 32), "Auto & Ancillaries": (14, 20, 26),
        "Chemicals": (18, 25, 33), "Building Materials": (24, 34, 46),
        "Logistics / Ports / Shipping": (16, 22, 30), "Diversified / Holding Co.": (12, 17, 24),
        "Hospitals": (38, 50, 64), "Hotels": (24, 34, 46), "Telecom": (12, 18, 28),
        "Real Estate": (16, 24, 34), "Metals & Mining (cyclical)": (8, 11, 15),
        "Energy / Oil & Gas": (8, 11, 14), "Utilities & Power": (11, 15, 19),
    }
    pe_lo, pe_base, pe_hi = _PE_BANDS.get(_bucket, (14, 20, 28))
    _heavy_bucket = _bucket in ("Metals & Mining (cyclical)", "Energy / Oil & Gas",
                                "Utilities & Power", "Real Estate")

    if is_financial and (bvps or (base_ni and base_ni > 0 and shares)):
        # Banks/NBFCs/insurers/financial holdcos: value on BOOK and normalised earnings power —
        # a lender is worth the BETTER of the two (protects a quality bank's franchise premium
        # AND a depressed/early-stage lender's book), never EV/EBITDA or a tiny trailing EPS.
        gj = min(max(g, 0.04), 0.09)
        eps_n = (base_ni / shares) if (base_ni and shares) else None
        roe_norm = (base_ni / equity) if (base_ni and equity) else roe_frac
        cands, hows = [], []
        if bvps and roe_frac and roe_frac > COE + 0.01 and COE > gj:
            jpb = (roe_frac - gj) / (COE - gj)
            cands.append(jpb * bvps); hows.append(f"justified P/B {jpb:.1f}×")
        if eps_n and eps_n > 0:
            fair_pe = min(max(12 + 0.4 * g * 100, 12), 18)   # Indian lenders: ~12–18× band
            cands.append(eps_n * fair_pe); hows.append(f"EPS ₹{eps_n:.0f}×{fair_pe:.0f}")
        if bvps:
            pbf = max(0.9, min(2.2, (roe_norm / COE) if (roe_norm and roe_norm > 0) else 1.0))
            cands.append(pbf * bvps); hows.append(f"{pbf:.1f}× book ₹{bvps:,.0f}")
        if cands:
            model_iv = max(cands)
            # A lender can't be reliably out-valued vs the market by a generic model, so anchor
            # to the market price (flag only clear divergences) — this reproduces the framework's
            # near-universal "hold, add on a dip" for financials instead of false add/trim signals.
            model_iv = max(0.80 * price, min(1.15 * price, model_iv))
            intrinsic = round(model_iv, 2)
            scen = {"bear": round(intrinsic * 0.80, 1), "base": round(intrinsic, 1),
                    "bull": round(intrinsic * 1.25, 1)}
            method = (f"P/B + normalised earnings, market-anchored (best of: {'; '.join(hows)}) "
                      f"— model for financials")
    elif _heavy_bucket and (ebitda_v or ebitda_hist) and shares:
        # Commodity/regulated cyclicals (metals, energy, utilities, realty): value the
        # enterprise on MID-CYCLE EBITDA at the firm's own revealed multiple (market-anchored),
        # not a fixed sector guess — that keeps quality's premium and avoids false bargains.
        ebitda_base = ebitda_v
        ebitda_state = "trailing"
        if ebitda_hist and len(ebitda_hist) >= 3:
            ebitda_med = _stats.median(ebitda_hist)
            if ebitda_v and ebitda_med and (ebitda_v < 0.6 * ebitda_med or ebitda_v > 1.6 * ebitda_med) and rev_intact:
                ebitda_base, ebitda_state = ebitda_med, "mid-cycle"
        cap = {"Utilities & Power": 14, "Energy / Oil & Gas": 10, "Metals & Mining (cyclical)": 9,
               "Real Estate": 16}.get(_bucket, 12)
        cur = evebitda if (evebitda and 3 < evebitda < cap * 1.4) else cap * 0.85
        tgt = min(max(cur, 4), cap)
        per_sh = lambda m: round((m * ebitda_base - (net_debt or 0)) / shares, 1)
        base_v = per_sh(tgt)
        intrinsic = base_v
        scen = {"bear": max(per_sh(max(tgt * 0.75, 3.5)), round(base_v * 0.6, 1)) if base_v > 0 else per_sh(max(tgt * 0.75, 3.5)),
                "base": base_v, "bull": per_sh(min(tgt * 1.25, cap + 1))}
        method = (f"EV/EBITDA · {tgt:.1f}× {ebitda_state} EBITDA (firm's own multiple) − net debt, "
                  f"per share — model for {_bucket}")
    elif base_ni and base_ni > 0 and shares:
        # Compounders / capital goods / pharma / auto / chemicals: normalised earnings power,
        # valued at a fair multiple ANCHORED to the market's own revealed multiple (forward P/E
        # preferred; it already reflects normalised earnings), disciplined by the sector band.
        base_eps = base_ni / shares
        mkt_mult = fwd_eps and price and (price / fwd_eps if fwd_eps > 0 else None)
        if not (mkt_mult and 5 <= mkt_mult <= 90):
            mkt_mult = pe if (pe and 6 <= pe <= 60) else None       # trailing, only if not distorted
        if mkt_mult:                          # lean on the market multiple, sector band as a light prior
            fair_pe = 0.60 * mkt_mult + 0.40 * pe_base
            fair_pe = max(mkt_mult * 0.75, min(mkt_mult * 1.25, fair_pe))
            anchor = "market-anchored"
        else:
            fair_pe = pe_base
            anchor = "sector band"
        lo_pe, hi_pe = fair_pe * 0.78, fair_pe * 1.28
        bull_eps = max(fwd_eps or 0, base_eps)   # Yahoo consensus forward EPS = the recovery leg
        intrinsic = round(base_eps * fair_pe, 2)
        scen = {"bear": round(0.85 * base_eps * lo_pe, 1), "base": round(base_eps * fair_pe, 1),
                "bull": round(bull_eps * hi_pe, 1)}
        method = (f"Normalised earnings · EPS ₹{base_eps:.1f} ({earn_state}) × fair P/E {fair_pe:.0f} "
                  f"({anchor}) — {_bucket or 'diversified'}")

    if intrinsic is None:                # last-resort earnings power / price
        if base_ni and base_ni > 0 and shares:
            base_eps = base_ni / shares
            intrinsic = round(base_eps * pe_base, 2)
            scen = {"bear": round(base_eps * pe_lo, 1), "base": round(base_eps * pe_base, 1),
                    "bull": round(base_eps * pe_hi, 1)}
            method = f"Normalised earnings · EPS ₹{base_eps:.1f} ({earn_state}) × fair P/E {pe_base:.0f}"
        else:
            intrinsic = price
            method = "Insufficient data — showing price"

    # Holding-company discount (framework §7.2 / §10.2): a holdco's consolidated earnings
    # over-count value the market never fully credits — apply a ~20% conglomerate discount.
    if _bucket == "Diversified / Holding Co." and intrinsic and "Insufficient" not in method:
        intrinsic = round(intrinsic * 0.80, 2)
        if scen:
            scen = {k: round(v * 0.80, 1) for k, v in scen.items()}
        method += " · −20% holdco discount"

    iv = {"intrinsic": intrinsic, "price": price, "method": method}
    if scen:
        iv["scen"] = scen
    mos = (intrinsic - price) / intrinsic * 100 if intrinsic else 0

    # ---------- reverse-DCF expectation test (what growth does the PRICE already assume?) ----------
    # Solve for the FCF/owner-earnings growth that makes a 2-stage DCF equal today's market cap.
    base_cf = fcf if (fcf and fcf > 0) else (ni if (ni and ni > 0) else None)
    if not is_financial and base_cf and mcap and mcap > 0:
        def _pv(gr):
            w, tg, f, pv = 0.11, 0.05, base_cf, 0.0
            for t in range(1, 6):
                f *= (1 + gr); pv += f / ((1 + w) ** t)
            return pv + (f * (1 + tg) / (w - tg)) / ((1 + w) ** 5)
        implied = None
        if _pv(0.40) >= mcap:               # solvable within a sane band
            lo, hi = -0.10, 0.40
            for _ in range(40):
                mid = (lo + hi) / 2
                if _pv(mid) < mcap: lo = mid
                else: hi = mid
            implied = (lo + hi) / 2
        if implied is None:
            cls, txt = "Unrealistic", "price implies >40%/yr FCF growth for a decade — priced for perfection"
        else:
            diff = implied - g
            cls = ("Conservative" if implied <= g else "Reasonable" if diff <= 0.03
                   else "Demanding" if diff <= 0.08 else "Unrealistic")
            txt = f"price implies ~{implied*100:.0f}%/yr FCF growth for a decade vs ~{g*100:.0f}% history"
        iv["impliedGrowth"] = round(implied * 100, 1) if implied is not None else None
        iv["expectation"] = cls
        iv["expectationText"] = txt
    elif is_financial:
        iv["expectation"] = "n/a"
        iv["expectationText"] = "reverse-DCF not meaningful for a lender — judged on P/B × sustainable ROE"

    # ---------- price-band engine (Section 17): at what price to add / accumulate / hold / trim ----------
    # Anchored on the scenario band (bear/base/bull) so a quality name trading between its
    # base and bull value reads HOLD (don't add) — never a mechanical TRIM. A TRIM signal
    # requires the price to sit materially ABOVE even the bull case (framework §7).
    if intrinsic and intrinsic > 0 and iv["method"] != "Insufficient data — showing price":
        base_iv = (scen["base"] if scen else intrinsic)
        bull_iv = (scen["bull"] if scen else intrinsic * 1.4)
        bear_iv = (scen["bear"] if scen else intrinsic * 0.65)
        bull_iv = max(bull_iv, base_iv * 1.15)         # keep the band sane if inputs are thin
        bear_iv = min(bear_iv, base_iv * 0.90)
        b = {
            "strongAdd": round(bear_iv, 1),             # deep value — bear-case worth
            "accumulate": round(base_iv * 0.85, 1),     # margin-of-safety buy band (10–15% below base)
            "fairLow": round(base_iv * 0.90, 1),
            "fairHigh": round(base_iv * 1.10, 1),
            "trim": round(bull_iv * 1.10, 1),           # materially above even the bull case
        }
        # Valuation richness is NOT a sell trigger for a quality holder (framework §7, §10.2:
        # never trim into a cyclical upturn). So a rich name is "hold, don't add"; an outright
        # "trim on valuation" flag is reserved for the truly extreme (>~30% above the bull case)
        # and is suppressed entirely for cyclical/commodity buckets.
        extreme = round(bull_iv * 1.30, 1)
        if price <= b["strongAdd"]:
            zone = "STRONG ADD"
        elif price <= b["accumulate"]:
            zone = "ACCUMULATE"
        elif price <= b["fairHigh"]:
            zone = "HOLD (fair value)"
        elif price <= b["trim"] or _heavy_bucket or is_financial or price <= extreme:
            zone = "HOLD (rich — don't add)"
        else:
            zone = "Expensive — trim only for portfolio reasons"
        iv["bands"] = b
        iv["zone"] = zone

    # ---------- enrich Key Ratios + Debt with more traditional/critical ratios ----------
    bvps = (equity / shares) if (equity and shares) else None
    ebitda_m = (ebitda_v / rev_now * 100) if (ebitda_v and rev_now) else None
    if dy: ratios.append({"k": "Dividend yield", "v": f"{dy:.2f}%"})
    if bvps: ratios.append({"k": "Book value/sh", "v": f"₹{bvps:,.0f}"})
    if ebitda_m is not None: ratios.append({"k": "EBITDA margin", "v": f"{ebitda_m:.1f}%"})
    if fcf and ni and ni > 0: ratios.append({"k": "FCF / PAT", "v": f"{fcf/ni:.2f}×"})
    if fcf and mcap: ratios.append({"k": "FCF yield", "v": f"{fcf/mcap*100:.1f}%"})
    if g: ratios.append({"k": "Rev CAGR (3y)", "v": f"{g*100:.0f}%"})
    if pe and g and g > 0: ratios.append({"k": "PEG", "v": f"{pe/(g*100):.2f}"})
    if icov: ratios.append({"k": "Interest cover", "v": f"{icov:.1f}×"})
    debt["netDebtEbitda"] = round(net_debt / ebitda_v, 2) if (net_debt is not None and ebitda_v and ebitda_v > 0) else None
    debt["cfoDebt"] = round(ocf / total_debt * 100) if (ocf and total_debt) else None
    debt["cashBal"] = cr(endcash)
    debt["fcfDebt"] = round(fcf / total_debt * 100) if (fcf and total_debt) else None

    # ---------- forensic heuristics ----------
    forensic = []
    if ocf and ni and ni != 0:
        ratio = ocf / ni
        if 0.8 <= ratio <= 1.4:
            forensic.append({"t": f"Cash flow backs profit (CFO/PAT {ratio:.2f}×) — clean", "s": "green"})
        elif ratio < 0.8:
            forensic.append({"t": f"CFO/PAT only {ratio:.2f}× — profit not fully cash-backed", "s": "amber"})
        else:
            forensic.append({"t": f"CFO/PAT {ratio:.2f}× — strong cash conversion", "s": "green"})
    if fcf is not None:
        if fcf < 0:
            forensic.append({"t": "Negative free cash flow — funded by debt/reserves", "s": "amber"})
        else:
            forensic.append({"t": "Positive free cash flow", "s": "green"})
    if de is not None:
        if de > 1.0:
            forensic.append({"t": f"High leverage (D/E {de:.2f}) — solvency risk in downturns", "s": "red"})
        elif de < 0.3:
            forensic.append({"t": f"Low leverage (D/E {de:.2f}) — sturdy balance sheet", "s": "green"})
        else:
            forensic.append({"t": f"Moderate leverage (D/E {de:.2f})", "s": "amber"})
    # receivables growing faster than revenue
    if recv_s is not None and rev_s is not None and len(recv_s) > 1 and len(rev_s) > 1:
        r0, r1 = val(recv_s, 0), val(recv_s, 1)
        s0, s1 = val(rev_s, 0), val(rev_s, 1)
        if r0 and r1 and s0 and s1 and r1 > 0 and s1 > 0:
            rg = r0 / r1 - 1; sg = s0 / s1 - 1
            if rg - sg > 0.15:
                forensic.append({"t": f"Receivables up {rg*100:.0f}% vs sales {sg*100:.0f}% — watch collections", "s": "amber"})
    if not forensic:
        forensic.append({"t": "Not enough statement history for forensic checks", "s": "amber"})

    # ---------- policy / plans (from available signals) ----------
    policy = []
    if div_s is not None and val(div_s):
        policy.append("Pays a dividend — returns cash to shareholders")
    if fcf and fcf > 0:
        policy.append("Self-funding: operations cover investment needs")
    else:
        policy.append("In investment/expansion phase — capex above cash generation")
    if de is not None and de < 0.3:
        policy.append("Conservative capital structure, low reliance on debt")
    elif de is not None:
        policy.append("Uses leverage — monitor refinancing and rates")
    if info.get("longBusinessSummary"):
        policy.append(info["longBusinessSummary"])

    # ---------- NSE shareholding pattern + promoter pledge ----------
    shareholding = pledge_info = None
    if resolved.endswith(".NS"):
        shareholding = nse.shareholding(base_sym)
        pledge_info = nse.pledge(base_sym)
        if shareholding is not None and pledge_info is not None:
            shareholding["pledgedPct"] = pledge_info.get("pledgedPct")

    # ---------- NSE corporate data (actions / board meetings / announcements) ----------
    corp_actions, board, announce = [], [], []
    if resolved.endswith(".NS"):
        ca = nse.corporate_actions(base_sym)
        if isinstance(ca, list):
            for x in ca[:6]:
                corp_actions.append({
                    "subject": x.get("subject") or x.get("purpose") or "—",
                    "exDate": x.get("exDate") or x.get("recDate") or "",
                })
        an = nse.announcements(base_sym)
        if isinstance(an, list):
            for x in an[:8]:
                url = x.get("attchmntFile") or ""
                announce.append({
                    "subject": (x.get("subject") or x.get("desc") or "—"),
                    "date": (x.get("an_dt") or x.get("sort_date") or "")[:20],
                    "url": url if url.startswith("http") else "",
                    "text": (x.get("attchmntText") or "").strip(),
                })

    # Screener financials were fetched once above (sfin_data); reuse for the forensic module.
    forensic_ratios = build_forensic(cf, bs, inc, info, endcash, shareholding, pledge_info, sfin_data)

    # ---------- decision scorecard (§15), confidence (§22), scenarios (§21) ----------
    fc = forensic_ratios.get("counts", {})
    fgood, fbad = fc.get("good", 0), fc.get("bad", 0)
    def clamp(v, lo, hi): return max(lo, min(hi, v))
    # Earnings quality & forensic integrity (15)
    s_eq = round(clamp(15 - fbad * 2, 0, 15))
    # Financial strength (10) — leverage & coverage
    s_fs = 10
    if de is not None:
        s_fs -= 5 if de > 2 else 3 if de > 1 else 1 if de > 0.5 else 0
    if icov is not None and icov < 3:
        s_fs -= 2
    s_fs = clamp(s_fs, 0, 10)
    # Capital efficiency (10) — ROCE / ROIC
    s_ce = 5
    if roce is not None:
        s_ce = 10 if roce > 18 else 8 if roce > 12 else 6 if roce > 9 else 3
    # Growth runway (10) — revenue CAGR proxy
    s_gr = round(clamp(g * 100 / 2, 0, 10))
    # Valuation / margin of safety (15)
    s_val = 7
    if mos is not None:
        s_val = 15 if mos > 25 else 11 if mos > 0 else 6 if mos > -30 else 3
    # Business quality & moat (15) — proxy from returns (refine via research)
    s_bq = 9
    if roe_calc is not None:
        s_bq = 14 if roe_calc > 20 else 11 if roe_calc > 14 else 9 if roe_calc > 9 else 6
    # Management & capital allocation (10) — proxy
    s_mg = 6
    if val(safe_row(cf, "Cash Dividends Paid", "Common Stock Dividend Paid")) or \
       val(safe_row(cf, "Repurchase Of Capital Stock")):
        s_mg += 1
    # Sector/policy (5), Geopolitical (5), Portfolio fit (5) — neutral proxies pending research
    s_sp, s_geo, s_pf = 3, 3, 3
    total = s_eq + s_fs + s_ce + s_gr + s_val + s_bq + s_mg + s_sp + s_geo + s_pf
    scorecard = {
        "total": total,
        "rows": [
            {"k": "Business quality & moat", "v": s_bq, "max": 15, "note": "proxy from returns — refine via research"},
            {"k": "Management & capital allocation", "v": s_mg, "max": 10, "note": "proxy"},
            {"k": "Earnings quality & forensic", "v": s_eq, "max": 15, "note": f"{fbad} red flag(s)"},
            {"k": "Financial strength", "v": s_fs, "max": 10, "note": "leverage & coverage"},
            {"k": "Capital efficiency", "v": s_ce, "max": 10, "note": "ROCE/ROIC"},
            {"k": "Growth runway & quality", "v": s_gr, "max": 10, "note": f"~{g*100:.0f}% rev CAGR"},
            {"k": "Valuation / margin of safety", "v": s_val, "max": 15, "note": iv.get("expectation", "")},
            {"k": "Sector / policy / macro", "v": s_sp, "max": 5, "note": "neutral — set by research"},
            {"k": "Geopolitical resilience", "v": s_geo, "max": 5, "note": "neutral — set by research"},
            {"k": "Portfolio fit & risk", "v": s_pf, "max": 5, "note": "neutral — set at portfolio level"},
        ],
        "note": "A decision AID, not a mechanical rule. Business-quality, management, policy & portfolio rows are proxies until research refines them.",
    }
    # Confidence (§22)
    missing = []
    if not quarters or len(quarters) < 4: missing.append("quarterly history")
    if intrinsic is None or "Insufficient" in method: missing.append("valuation")
    if not shareholding: missing.append("shareholding/governance")
    if not forensic_ratios.get("categories"): missing.append("forensics")
    confidence = "HIGH" if not missing else "MEDIUM" if len(missing) == 1 else "LOW"
    confidence_note = ("Current data covers the critical variables." if not missing
                       else "Missing/limited: " + ", ".join(missing) + " — plus qualitative depth needs research.")
    # Scenarios (§21) — use the sector-appropriate bear/base/bull band when available
    scenarios = None
    if intrinsic and intrinsic > 0 and "Insufficient" not in method:
        if scen:
            scenarios = {
                "bull": scen["bull"], "base": scen["base"], "bear": scen["bear"],
                "severe": round(scen["bear"] * 0.65, 1),
            }
        else:
            scenarios = {
                "bull": round(intrinsic * 1.5, 1), "base": round(intrinsic, 1),
                "bear": round(intrinsic * 0.65, 1), "severe": round(intrinsic * 0.40, 1),
            }

    # ---------- dynamic FRAMEWORK READ (Portfolio_Framework.docx §7/§9) ----------
    # The framework AS A FEATURE: verdict · conviction 1-5 · basis · long-term qualification ·
    # MoS buy band · plain-language buy note — all recomputed from LIVE data on every refresh,
    # so they update automatically each quarter as new results are filed (vs the dated Excel).
    framework_read = None
    if intrinsic and "Insufficient" not in method and iv.get("bands"):
        b = iv["bands"]; z = iv.get("zone", "")
        # Conviction 1-5 from the 100-pt scorecard (which already folds in forensic quality) —
        # NOT the raw forensic "bad" count, which is a screen and over-fires on healthy names.
        tot = scorecard["total"]
        conv = 5 if tot >= 75 else 4 if tot >= 60 else 3 if tot >= 48 else 2 if tot >= 36 else 1
        # Long-term qualification (§7.3), from sector & through-cycle returns. A CONFIRMED
        # accounting/governance red flag (a MUST-SELL) comes from research/news, not this screen.
        roe_n = (base_ni / equity * 100) if (base_ni and equity) else roe_calc
        if _heavy_bucket or earn_state != "trailing":
            ltq = "YES (cyclical / recovery) — qualifies, but manage through the cycle"
        elif (roe_n or 0) >= 15:
            ltq = "YES — franchise & returns support a decade of ownership"
        elif (roe_n or 0) >= 9:
            ltq = "YES (watch) — qualifies; monitor returns & execution"
        else:
            ltq = "CONDITIONAL — sub-cost-of-capital returns; needs a turnaround thesis"
        band_str = f"₹{b['accumulate']:,.0f}–{b['fairLow']:,.0f}"
        if z.startswith("STRONG ADD"):
            verdict, note = "Add", f"Below bear-case worth — deploy fresh capital now; MoS buy band {band_str}."
        elif z.startswith("ACCUMULATE"):
            verdict, note = "Accumulate", f"Buy in tranches on weakness toward the MoS band {band_str}."
        elif "fair value" in z:
            verdict, note = "Hold", f"Around fair value — hold; add only on a dip into {band_str}."
        elif "rich" in z:
            verdict, note = "Hold", f"Richly valued — hold, don't add above ₹{b['fairHigh']:,.0f}; re-add in {band_str}."
        else:
            verdict, note = "Hold", "Rich vs the bull case — hold; trim only for portfolio / discipline reasons."
        basis = ("DD (live)" if (quarters and len(quarters) >= 4 and forensic_ratios.get("categories"))
                 else "P (live · partial data)")
        framework_read = {
            "verdict": verdict, "conviction": conv, "basis": basis, "ltQualify": ltq,
            "mosBuyBand": band_str, "buyNote": note,
            "fairIV": (scen["base"] if scen else intrinsic),
            "bearIV": (scen["bear"] if scen else None), "bullIV": (scen["bull"] if scen else None),
            "expectation": iv.get("expectation"),
            "note": "Quant read from live data — a confirmed forensic/policy red flag (from research) overrides to Exit.",
        }

    # ---------- bank / NBFC / FI view (Issue 2): the RIGHT metrics for a lender ----------
    # EBITDA/current ratio/inventory days are meaningless for lenders. Show asset quality
    # (GNPA/NNPA), the funding franchise (deposits), the income split (interest vs fee), NII,
    # RoA/RoE — sourced live from Screener; metrics that live only in bank filings are flagged.
    bank_view = None
    if is_financial:
        try:
            sqb = screener_web.fetch_quarterly(base_sym)   # cached; carries gnpa/nnpa/nii
        except Exception:
            sqb = None
        emap = {6: "Q1", 9: "Q2", 12: "Q3", 3: "Q4"}
        def _flabel(mo, yr):
            fy = (yr + 1) if mo >= 4 else yr
            return f"{emap.get(mo, '')} FY{fy % 100:02d}"
        gnpa = nnpa = npa_asof = None
        nii_series = []
        # Asset quality: the standalone quarters table (regulatory disclosure) is authoritative.
        try:
            npa = screener_web.fetch_bank_npa(base_sym)
        except Exception:
            npa = None
        if npa:
            gnpa, nnpa = npa.get("gnpa"), npa.get("nnpa")
            _m = re.match(r"([A-Za-z]{3})\s+(\d{4})", npa.get("asOf") or "")
            if _m:
                _mo = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}.get(_m.group(1).lower())
                npa_asof = _flabel(_mo, int(_m.group(2))) if _mo else npa.get("asOf")
        if sqb:
            if gnpa is None:                                # fallback: consolidated (last non-null)
                for x in sqb:
                    if x.get("gnpa") is not None:
                        gnpa, npa_asof = x["gnpa"], _flabel(x["m"], x["y"])
                    if x.get("nnpa") is not None:
                        nnpa = x["nnpa"]
            for x in sqb[-8:]:
                lbl = f"{emap.get(x['m'],'')} {'FY%02d' % ((x['y']+1)%100) if x['m']>=4 else 'FY%02d' % (x['y']%100)}"
                nii_series.append({"q": lbl, "nii": x.get("nii"), "otherIncome": x.get("otherIncome"),
                                   "np": x.get("np"), "gnpa": x.get("gnpa"), "nnpa": x.get("nnpa"),
                                   "intIncome": x.get("rev"), "intExp": x.get("interest")})
        # per-quarter NII margin for the trend
        for x in nii_series:
            x["niiMargin"] = round(x["nii"] / x["intIncome"] * 100, 1) if (x.get("nii") and x.get("intIncome")) else None
        latest = sqb[-1] if sqb else {}
        total_assets = val(safe_row(bs, "Total Assets"))
        deposits_cr = (sfin_data or {}).get("deposits")
        borrow_cr = (sfin_data or {}).get("borrowings")
        nii_l, oi_l, ie_l = latest.get("nii"), latest.get("otherIncome"), latest.get("interest")
        funding = (deposits_cr or 0) + (borrow_cr or 0)
        total_assets_cr = (total_assets / 1e7) if total_assets else None   # raw ₹ -> ₹ Cr
        nim_est = round(nii_l * 4 / total_assets_cr * 100, 2) if (nii_l and total_assets_cr) else None
        cof_est = round(ie_l * 4 / funding * 100, 2) if (ie_l and funding) else None
        roa_v = round(ni / total_assets * 100, 2) if (ni and total_assets) else None
        # NII YoY (vs 4 quarters ago)
        nii_yoy = None
        _s = [x for x in nii_series if x.get("nii") is not None]
        if len(_s) >= 5 and _s[-5]["nii"]:
            nii_yoy = round((_s[-1]["nii"] / _s[-5]["nii"] - 1) * 100, 1)
        bank_view = {
            "isNBFC": _bucket == "NBFCs & Financiers",
            "bucket": _bucket,
            "gnpa": gnpa, "nnpa": nnpa, "npaAsOf": npa_asof,
            "depositsCr": round(deposits_cr) if deposits_cr else None,
            "intIncomeCr": latest.get("rev"), "intExpCr": ie_l,
            "niiCr": nii_l, "otherIncomeCr": oi_l,
            "feeShare": round(oi_l / (nii_l + oi_l) * 100) if (nii_l and oi_l) else None,
            "roa": roa_v, "roe": round(roe_calc, 1) if roe_calc is not None else None,
            "niiMargin": round(nii_l / latest.get("rev") * 100, 1) if (nii_l and latest.get("rev")) else None,
            "nimEst": nim_est, "cofEst": cof_est, "niiYoY": nii_yoy,
            "series": nii_series,
            # honest gaps — these live only in the bank's results filing / investor deck
            "fromFilings": ["NIM (exact, on avg earning assets)", "CASA %", "PCR / coverage",
                            "Cost-to-income (ex-provisions)", "Credit cost & slippage", "CET1 / CRAR", "LCR"],
        }
        # ---- Bank-specific Key Ratios (override the generic industrial set) ----
        bvps_fin = (equity / shares) if (equity and shares) else None
        bank_ratios = []
        if roa_v is not None: bank_ratios.append({"k": "RoA", "v": f"{roa_v:.2f}%"})
        if roe_calc is not None: bank_ratios.append({"k": "RoE", "v": f"{roe_calc:.1f}%"})
        if nim_est is not None: bank_ratios.append({"k": "NIM (est.)", "v": f"{nim_est:.2f}%"})
        if cof_est is not None: bank_ratios.append({"k": "Cost of funds (est.)", "v": f"{cof_est:.2f}%"})
        if gnpa is not None: bank_ratios.append({"k": "Gross NPA", "v": f"{gnpa}%"})
        if nnpa is not None: bank_ratios.append({"k": "Net NPA", "v": f"{nnpa}%"})
        if bank_view["feeShare"] is not None: bank_ratios.append({"k": "Fee/other income", "v": f"{bank_view['feeShare']}% of income"})
        if nii_yoy is not None: bank_ratios.append({"k": "NII growth (YoY)", "v": f"{nii_yoy:+.1f}%"})
        if pb: bank_ratios.append({"k": "P/B", "v": f"{pb:.1f}"})
        if bvps_fin: bank_ratios.append({"k": "Book value/sh", "v": f"₹{bvps_fin:,.0f}"})
        if pe: bank_ratios.append({"k": "P/E", "v": f"{pe:.1f}"})
        if dy: bank_ratios.append({"k": "Dividend yield", "v": f"{dy:.2f}%"})
        if bank_ratios:
            ratios = bank_ratios         # replace the industrial ratios entirely for lenders

    company = {
        "name": nse_name or info.get("longName") or NAME_BY_SYM.get(base_sym, base_sym),
        "ticker": base_sym, "bse": base_sym, "nse": base_sym,
        "bankView": bank_view,
        "frameworkRead": framework_read,
        "sector": nse_sector or info.get("sector") or info.get("industry") or "—",
        "exchange": exch, "resolved": resolved, "priceSource": price_source,
        "price": price, "prevClose": prev, "currency": "₹",
        "mcap": cr(mcap) if mcap else None,
        "pe": round(pe, 1) if pe else "—", "pb": round(pb, 1) if pb else "—",
        "divYield": round(dy, 2) if dy else 0,
        "high52": nse_hi or info.get("fiftyTwoWeekHigh") or price,
        "low52": nse_lo or info.get("fiftyTwoWeekLow") or price,
        "kpis": kpis, "sources": sources, "uses": uses,
        "capex": capex_rows, "fcf": fcf_rows, "wc": wc, "ccc": ccc_rows,
        "quarters": quarters, "iv": iv, "debt": debt,
        "forensic": forensic, "ratios": ratios, "policy": policy,
        "forensicRatios": forensic_ratios,
        "scorecard": scorecard, "confidence": confidence, "confidenceNote": confidence_note,
        "scenarios": scenarios,
        "sectorLens": sectors.profile(nse_sector or info.get("sector") or "", f"{nse_name or info.get('longName') or ''} {base_sym}"),
        "shareholding": shareholding,
        "corpActions": corp_actions, "boardMeetings": board, "announcements": announce,
        "personas": build_personas(info, mos, de, roe_calc, pe, npm, fcf, ocf, ni),
        "asOf": dt.date.today().isoformat(),
    }
    return company


def build_forensic(cf, bs, inc, info, endcash, shareholding=None, pledge_info=None, sfin=None):
    """
    Compute the forensic / diagnostic ratio set (from the user's framework),
    grouped into clickable categories, each ratio tagged good / watch / bad / na
    with a plain-language interpretation. India-specific governance items that
    live only in annual reports / NSE filings are surfaced as a 'check filings'
    category rather than fabricated.
    """
    def A(df, names, i=0):
        return val(safe_row(df, *names), i)

    # ---- raw inputs (latest = index 0, prior year = index 1) ----
    ocf = A(cf, ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities"])
    capex = A(cf, ["Capital Expenditure"]); capex = abs(capex) if capex else None
    fcf = A(cf, ["Free Cash Flow"])
    if fcf is None and ocf is not None and capex is not None:
        fcf = ocf - capex
    ni = A(inc, ["Net Income"]); ni1 = A(inc, ["Net Income"], 1)
    rev = A(inc, ["Total Revenue", "Operating Revenue"]); rev1 = A(inc, ["Total Revenue", "Operating Revenue"], 1)
    cogs = A(inc, ["Cost Of Revenue", "Reconciled Cost Of Revenue"])
    ebitda = A(inc, ["EBITDA", "Normalized EBITDA"]); ebitda1 = A(inc, ["EBITDA", "Normalized EBITDA"], 1)
    ebit = A(inc, ["EBIT"]); ebit1 = A(inc, ["EBIT"], 1)
    intexp = A(inc, ["Interest Expense"])
    taxrate = A(inc, ["Tax Rate For Calcs"]); taxrate1 = A(inc, ["Tax Rate For Calcs"], 1)
    dep = A(inc, ["Reconciled Depreciation"]) or A(cf, ["Depreciation And Amortization", "Depreciation Amortization Depletion"])
    assets = A(bs, ["Total Assets"]); assets1 = A(bs, ["Total Assets"], 1)
    avg_assets = ((assets + assets1) / 2) if (assets and assets1) else assets
    equity = A(bs, ["Stockholders Equity", "Common Stock Equity"])
    total_debt = A(bs, ["Total Debt"]) or info.get("totalDebt")
    total_debt1 = A(bs, ["Total Debt"], 1)
    net_debt = A(bs, ["Net Debt"])
    if net_debt is None and total_debt is not None and endcash is not None:
        net_debt = total_debt - endcash
    invcap = A(bs, ["Invested Capital"]); invcap1 = A(bs, ["Invested Capital"], 1)
    wc = A(bs, ["Working Capital"]); wc1 = A(bs, ["Working Capital"], 1)
    recv = A(bs, ["Receivables", "Accounts Receivable"]); recv1 = A(bs, ["Receivables", "Accounts Receivable"], 1)
    inv = A(bs, ["Inventory"]); inv1 = A(bs, ["Inventory"], 1)
    pay = A(bs, ["Payables", "Accounts Payable"])
    curdebt = A(bs, ["Current Debt", "Current Debt And Capital Lease Obligation"])
    goodwill = A(bs, ["Goodwill"])
    intang = A(bs, ["Goodwill And Other Intangible Assets", "Other Intangible Assets"])
    netppe = A(bs, ["Net PPE"]); grossppe = A(bs, ["Gross PPE"])
    intpaid = A(cf, ["Interest Paid Cff"]); intpaid = abs(intpaid) if intpaid else None
    divpaid = A(cf, ["Cash Dividends Paid", "Common Stock Dividend Paid"]); divpaid = abs(divpaid) if divpaid else None
    buyback = A(cf, ["Repurchase Of Capital Stock"]); buyback = abs(buyback) if buyback else 0.0
    acq = A(cf, ["Purchase Of Business", "Net Business Purchase And Sale"]); acq = abs(acq) if acq else 0.0

    # ---- Screener fallbacks (stated numbers only) for values the primary feed lacks ----
    sf = sfin or {}
    def sfv(k):
        v = sf.get(k)
        return v if isinstance(v, (int, float)) else None
    if total_debt is None and sfv("borrowings") is not None:
        total_debt = sfv("borrowings")
    if net_debt is None and total_debt is not None:
        net_debt = total_debt - (endcash or 0)
    if equity is None and sfv("netWorth") is not None:
        equity = sfv("netWorth")
    if invcap is None and sfv("netWorth") is not None and sfv("borrowings") is not None:
        invcap = sfv("netWorth") + sfv("borrowings")   # invested capital = equity + debt (standard)
    if ocf is None and sfv("cfo") is not None:
        ocf = sfv("cfo")

    WACC = 11.0  # %
    nopat = ebit * (1 - taxrate) if (ebit is not None and taxrate is not None) else (ebit * 0.75 if ebit else None)
    nopat1 = ebit1 * (1 - (taxrate1 if taxrate1 is not None else (taxrate or 0.25))) if ebit1 is not None else None
    roic = (nopat / invcap * 100) if (nopat and invcap) else None

    # ---- formatters ----
    fx = lambda v: f"{v:.2f}×"
    fp = lambda v: f"{v:.1f}%"
    fpp = lambda v: f"{v:+.1f} pp"
    fd = lambda v: f"{round(v)} d"
    fyr = lambda v: f"{v:.1f} yr"

    def hi(v, good, bad):   # higher is better
        if v is None: return None
        return "good" if v >= good else ("bad" if v < bad else "warn")

    def lo(v, good, bad):   # lower is better
        if v is None: return None
        return "good" if v <= good else ("bad" if v > bad else "warn")

    def band(v, lo_ok, hi_ok):  # inside band = good
        if v is None: return None
        return "good" if lo_ok <= v <= hi_ok else "warn"

    def safediv(a, b):
        return (a / b) if (a is not None and b not in (None, 0)) else None

    cats = []

    def C(name, desc=""):
        c = {"name": name, "desc": desc, "ratios": []}
        cats.append(c)
        return c

    def add(cat, n, formula, value, fmt, sig, msgs, src=None):
        if value is None or (isinstance(value, float) and math.isnan(value)):
            cat["ratios"].append({"n": n, "f": formula, "v": "—", "signal": "na",
                                  "interp": "Not available from the automated feed."})
            return
        txt = {"good": msgs[0], "warn": msgs[1], "bad": msgs[2]}.get(sig, "")
        if src:
            txt = (txt + f" · {src}").strip(" ·")
        cat["ratios"].append({"n": n, "f": formula, "v": fmt(value), "signal": sig or "na", "interp": txt})

    def manual(cat, n, why):
        cat["ratios"].append({"n": n, "f": "from filings", "v": "check", "signal": "na", "interp": why})

    # ===== 1. Earnings Quality =====
    c = C("Earnings Quality", "Are the reported profits actually real cash?")
    v = safediv(ocf, ni)
    add(c, "CFO / PAT", "Operating cash flow ÷ net profit", v, fx, hi(v, 1.0, 0.8) if v is not None else None,
        ("Profits are well backed by operating cash.", "Cash lags profit somewhat — monitor.",
         "Profit is not converting to cash — quality concern."))
    v = safediv(fcf, ni)
    add(c, "FCF / PAT", "Free cash flow ÷ net profit", v, fx, hi(v, 0.8, 0.4) if v is not None else None,
        ("Strong free cash after maintaining the business.", "Moderate conversion to distributable cash.",
         "Little profit survives as free cash — investigate."))
    v = safediv((ni - ocf), avg_assets) * 100 if (ni is not None and ocf is not None and avg_assets) else None
    add(c, "Accrual Ratio", "(PAT − CFO) ÷ avg assets", v, fp, lo(v, 0, 10) if v is not None else None,
        ("Earnings are cash-based, low accruals.", "Some accrual dependence.",
         "Accrual-heavy earnings — earnings-quality risk."))
    v = safediv(ocf, ebitda)
    add(c, "CFO / EBITDA", "Operating cash flow ÷ EBITDA", v, fx, hi(v, 0.8, 0.6) if v is not None else None,
        ("EBITDA converts well into cash.", "Partial EBITDA-to-cash conversion.",
         "EBITDA isn't turning into cash — warning."))
    v = safediv(ocf, rev) * 100 if (ocf is not None and rev) else None
    add(c, "CFO / Sales", "Operating cash flow ÷ revenue", v, fp, hi(v, 12, 6) if v is not None else None,
        ("Healthy cash generated per rupee of sales.", "Modest cash margin.",
         "Thin operating cash margin."))

    # ===== 2. Working-Capital Forensics =====
    c = C("Working-Capital Forensics", "Is growth real, or is cash getting trapped?")
    dso = safediv(recv, rev) * 365 if (recv is not None and rev) else None
    dso_src = None
    if dso is None and sfv("dso") is not None:
        dso, dso_src = sfv("dso"), "via Screener"
    add(c, "DSO", "Receivables ÷ revenue × 365", dso, fd, lo(dso, 60, 120) if dso is not None else None,
        ("Customers pay quickly.", "Moderate collection period.", "Slow collections — receivable risk."), dso_src)
    dio = safediv(inv, cogs) * 365 if (inv is not None and cogs) else None
    dio_src = None
    if dio is None and sfv("dio") is not None:
        dio, dio_src = sfv("dio"), "via Screener"
    add(c, "DIO", "Inventory ÷ COGS × 365", dio, fd, lo(dio, 60, 120) if dio is not None else None,
        ("Inventory turns efficiently.", "Moderate inventory holding.", "Inventory sits long — possible stress."), dio_src)
    dpo = safediv(pay, cogs) * 365 if (pay is not None and cogs) else None
    dpo_src = None
    if dpo is None and sfv("dpo") is not None:
        dpo, dpo_src = sfv("dpo"), "via Screener"
    add(c, "DPO", "Payables ÷ COGS × 365", dpo, fd, ("good" if (dpo and 30 <= dpo <= 120) else "warn") if dpo is not None else None,
        ("Reasonable supplier financing.", "Unusual payment terms — check liquidity.", ""), dpo_src)
    ccc = (dso + dio - dpo) if (dso is not None and dio is not None and dpo is not None) else None
    ccc_src = "via Screener" if (dso_src or dio_src or dpo_src) else None
    if ccc is None and sfv("ccc") is not None:
        ccc, ccc_src = sfv("ccc"), "via Screener"
    add(c, "Cash Conversion Cycle", "DSO + DIO − DPO", ccc, fd, lo(ccc, 30, 90) if ccc is not None else None,
        ("Cash is freed quickly (low/negative cycle).", "Moderate cash lock-up.", "Lots of cash trapped in operations."), ccc_src)
    wcs = safediv(wc, rev) * 100 if (wc is not None and rev) else None
    wcs_src = None
    if wcs is None and sfv("wcDays") is not None:
        wcs, wcs_src = sfv("wcDays") / 365 * 100, "via Screener (WC days)"
    add(c, "Working Capital / Sales", "Operating WC ÷ revenue", wcs, fp, lo(wcs, 15, 35) if wcs is not None else None,
        ("Scalable — little capital needed for sales.", "Some working-capital intensity.",
         "Growth ties up a lot of working capital."), wcs_src)
    iwc = safediv((wc - wc1), (rev - rev1)) if (wc is not None and wc1 is not None and rev and rev1 and rev != rev1) else None
    if iwc is not None and (rev - rev1) < 0:
        iwc = None
    add(c, "Incremental WC / Δ Sales", "ΔWC ÷ Δrevenue", (iwc * 100 if iwc is not None else None), fp,
        lo(iwc * 100, 20, 50) if iwc is not None else None,
        ("Each new rupee of sales needs little extra cash.", "Growth is moderately cash-hungry.",
         "Growth is very cash-intensive."))
    rgp = ((recv / recv1 - 1) - (rev / rev1 - 1)) * 100 if (recv and recv1 and rev and rev1) else None
    add(c, "Receivable Growth Premium", "AR growth − revenue growth", rgp, fpp, lo(rgp, 0, 10) if rgp is not None else None,
        ("Receivables in line with sales — clean.", "Receivables outpacing sales a little.",
         "Receivables far outpacing sales — revenue-quality flag."))
    igp = ((inv / inv1 - 1) - (rev / rev1 - 1)) * 100 if (inv and inv1 and rev and rev1) else None
    add(c, "Inventory Growth Premium", "Inv growth − sales growth", igp, fpp, lo(igp, 0, 10) if igp is not None else None,
        ("Inventory in line with sales.", "Inventory building faster than sales.",
         "Inventory build-up — demand/obsolescence risk."))
    v = safediv(recv, ocf)
    add(c, "Receivables / CFO", "Receivables ÷ operating cash flow", v, fx, lo(v, 1, 2) if v is not None else None,
        ("Low reliance on uncollected receivables.", "Moderate receivable dependence.",
         "Heavy dependence on collecting receivables."))

    # ===== 3. Capital Efficiency =====
    c = C("Capital Efficiency", "Is growth actually profitable? (the real moat test)")
    add(c, "ROIC", "NOPAT ÷ invested capital", roic, fp, hi(roic, WACC + 3, WACC) if roic is not None else None,
        ("Returns comfortably exceed the cost of capital.", "Returns roughly match cost of capital.",
         "Returns below cost of capital — value destruction."))
    spread = (roic - WACC) if roic is not None else None
    add(c, "ROIC − WACC Spread", f"ROIC − WACC ({WACC:.0f}%)", spread, fpp, hi(spread, 0, -2) if spread is not None else None,
        ("Positive economic spread — creating value.", "Roughly breakeven on capital.",
         "Negative spread — destroying economic value."))
    iroic = safediv((nopat - nopat1), (invcap - invcap1)) * 100 if (nopat is not None and nopat1 is not None and invcap and invcap1 and abs(invcap - invcap1) > 0.02 * abs(invcap)) else None
    add(c, "Incremental ROIC", "ΔNOPAT ÷ Δinvested capital", iroic, fp, hi(iroic, WACC, 0) if iroic is not None else None,
        ("New capital is being deployed at high returns.", "New capital earns modest returns.",
         "New capital is being deployed poorly."))
    at = safediv(rev, avg_assets)
    add(c, "Asset Turnover", "Revenue ÷ avg assets", at, fx, hi(at, 1.0, 0.4) if at is not None else None,
        ("Assets are used productively.", "Moderate asset productivity.", "Low asset productivity (capital-heavy)."))
    fat = safediv(rev, netppe)
    add(c, "Fixed Asset Turnover", "Revenue ÷ net PP&E", fat, fx, hi(fat, 2.5, 1.0) if fat is not None else None,
        ("High output per rupee of fixed assets.", "Moderate fixed-asset productivity.",
         "Low fixed-asset productivity — possible over-investment."))
    ci = safediv(netppe, rev)
    add(c, "Capital Intensity", "Net PP&E ÷ revenue", ci, fx, lo(ci, 0.5, 1.0) if ci is not None else None,
        ("Asset-light and scalable.", "Moderately capital-intensive.", "Very capital-intensive business."))
    fcfroic = safediv(fcf, invcap) * 100 if (fcf is not None and invcap) else None
    add(c, "FCF Return on Capital", "FCF ÷ invested capital", fcfroic, fp, hi(fcfroic, 8, 0) if fcfroic is not None else None,
        ("Strong cash return on capital employed.", "Modest cash return on capital.",
         "Weak/negative cash return on capital."))

    # ===== 4. Debt & Financial Stress =====
    c = C("Debt & Financial Stress", "Can the business service and repay its debt from cash?")
    v = safediv(net_debt, ebitda)
    add(c, "Net Debt / EBITDA", "Net debt ÷ EBITDA", v, fx, lo(v, 1.0, 3.0) if v is not None else None,
        ("Low leverage relative to earnings.", "Moderate leverage.", "High leverage — balance-sheet risk."))
    v = safediv(ocf, total_debt) * 100 if (ocf is not None and total_debt) else (999 if (total_debt in (0, None) and ocf) else None)
    add(c, "CFO / Debt", "Operating cash flow ÷ total debt", v, fp, hi(v, 40, 20) if v is not None else None,
        ("Operations easily cover the debt.", "Adequate cash debt-coverage.", "Weak cash cover for debt."))
    v = safediv(fcf, total_debt) * 100 if (fcf is not None and total_debt) else (999 if (total_debt in (0, None) and fcf) else None)
    add(c, "FCF / Debt", "Free cash flow ÷ total debt", v, fp, hi(v, 25, 10) if v is not None else None,
        ("Strong free-cash debt repayment capacity.", "Moderate repayment capacity.", "Weak repayment capacity."))
    v = safediv(ocf, intpaid)
    add(c, "Cash Interest Coverage", "CFO ÷ interest paid", v, fx, hi(v, 5, 2) if v is not None else None,
        ("Interest easily covered by cash.", "Adequate cash interest cover.", "Genuine debt-service stress."))
    v = safediv(ebit, intexp)
    add(c, "EBIT / Interest", "EBIT ÷ interest expense", v, fx, hi(v, 4, 2) if v is not None else None,
        ("Comfortable accounting interest cover.", "Adequate interest cover.", "Thin interest cover — investigate."))
    v = safediv(net_debt, fcf) if (net_debt is not None and fcf and fcf > 0) else (0 if (net_debt is not None and net_debt <= 0) else None)
    add(c, "Net Debt / FCF", "Net debt ÷ free cash flow", v, fyr, lo(v, 2, 5) if v is not None else None,
        ("Net cash or quick deleveraging.", "Manageable years to repay.", "Slow, difficult deleveraging."))
    v = safediv(curdebt, ocf)
    add(c, "Short-Term Debt / CFO", "Current debt ÷ CFO", v, fx, lo(v, 0.5, 1.0) if v is not None else None,
        ("Low near-term refinancing need.", "Some refinancing dependence.", "High refinancing/liquidity risk."))
    v = ((total_debt / total_debt1 - 1) - (ebitda / ebitda1 - 1)) * 100 if (total_debt and total_debt1 and ebitda and ebitda1) else None
    add(c, "Debt Growth − EBITDA Growth", "Δdebt − ΔEBITDA", v, fpp, lo(v, 0, 10) if v is not None else None,
        ("Earnings growing faster than debt.", "Debt creeping up vs earnings.", "Leverage outrunning earnings — warning."))

    # ===== 5. Hidden Balance-Sheet & Accounting Risk =====
    c = C("Hidden Balance-Sheet Risk", "What's lurking in intangibles, capex and assets?")
    v = safediv(goodwill, equity) * 100 if (goodwill is not None and equity) else (0 if goodwill in (0, None) else None)
    add(c, "Goodwill / Equity", "Goodwill ÷ net worth", v, fp, lo(v, 10, 35) if v is not None else None,
        ("Little acquisition goodwill on the books.", "Moderate goodwill exposure.",
         "High goodwill — impairment risk."))
    v = safediv(intang, equity) * 100 if (intang is not None and equity) else None
    add(c, "Intangibles / Equity", "Intangibles ÷ net worth", v, fp, lo(v, 25, 60) if v is not None else None,
        ("Strong tangible capital backing.", "Some intangible-heavy capital.", "Weak tangible backing of equity."))
    v = safediv(capex, dep)
    add(c, "Capex / Depreciation", "Capex ÷ depreciation", v, fx, ("good" if (v and v >= 1) else ("warn" if (v and v >= 0.7) else "bad")) if v is not None else None,
        ("Asset base is being replenished/grown.", "Investment roughly matches wear.",
         "Persistent under-investment in assets."))
    v = safediv(dep, grossppe) * 100 if (dep is not None and grossppe) else None
    add(c, "Depreciation / Gross PP&E", "Depreciation ÷ gross PP&E", v, fp, band(v, 4, 12) if v is not None else None,
        ("Depreciation intensity looks normal.", "Depreciation assumptions worth a check.", ""))
    v = safediv(capex, ocf) * 100 if (capex is not None and ocf) else None
    add(c, "Capex / CFO", "Capex ÷ operating cash flow", v, fp, lo(v, 50, 90) if v is not None else None,
        ("Operations comfortably fund capex.", "Capex consumes much of operating cash.",
         "Capex eats almost all operating cash — low FCF."))

    # ===== 6. Management Behaviour =====
    c = C("Management Behaviour", "What are capital-allocation choices telling you?")
    v = safediv(divpaid, fcf) * 100 if (divpaid is not None and fcf and fcf > 0) else None
    add(c, "Dividend / FCF", "Dividends ÷ free cash flow", v, fp, lo(v, 60, 100) if v is not None else None,
        ("Dividends comfortably funded by cash.", "Dividend uses most free cash.",
         "Dividends exceed free cash — funded by debt/reserves."))
    v = safediv(buyback, fcf) * 100 if (fcf and fcf > 0) else None
    add(c, "Buybacks / FCF", "Buybacks ÷ free cash flow", v, fp, lo(v, 50, 100) if v is not None else None,
        ("Buybacks well within free cash.", "Buybacks use much of free cash.",
         "Buybacks exceed free cash — may need financing."))
    v = safediv(acq, ocf) * 100 if (ocf) else None
    add(c, "Acquisition Spend / CFO", "Acquisition outflow ÷ CFO", v, fp, lo(v, 10, 40) if v is not None else None,
        ("Growth is largely organic.", "Some reliance on acquisitions.",
         "Growth heavily acquisition-led — integration risk."))
    add(c, "Effective Tax Rate", "Tax ÷ pre-tax profit", (taxrate * 100 if taxrate is not None else None), fp,
        band(taxrate * 100, 15, 35) if taxrate is not None else None,
        ("Normal tax rate — no obvious red flag.", "Unusual tax rate — worth understanding why.", ""))

    # ===== 7. India Governance & Filings =====
    c = C("India Governance", "Promoter/pledge now pulled live from NSE; the rest need annual reports.")
    # -- Promoter pledge (live from NSE) --
    pledged = (pledge_info or {}).get("pledgedPct") if pledge_info else (shareholding or {}).get("pledgedPct")
    if pledged is not None:
        sig = "good" if pledged < 1 else ("warn" if pledged < 10 else "bad")
        add(c, "Promoter Pledge %", "pledged ÷ promoter shares (NSE)", pledged, fp, sig,
            ("Negligible pledging — no promoter stress signal.",
             "Some promoter pledging — monitor each quarter.",
             "High promoter pledge — major financial-stress red flag."))
    else:
        manual(c, "Promoter Pledge %", "Rising pledged promoter shares = stress signal. Check NSE shareholding pattern.")
    # -- Promoter holding trend (live from NSE) --
    trend = (shareholding or {}).get("promoterTrend") or []
    if len(trend) >= 2:
        delta = trend[-1]["promoter"] - trend[0]["promoter"]
        sig = "good" if delta >= -0.5 else ("warn" if delta >= -3 else "bad")
        c["ratios"].append({"n": "Promoter Holding Trend", "f": f"{trend[0]['date']} → {trend[-1]['date']} (NSE)",
                            "v": f"{trend[-1]['promoter']:.1f}% ({delta:+.1f})", "signal": sig,
                            "interp": ("Promoters holding steady/adding — aligned with shareholders." if sig == "good"
                                       else "Promoter stake slipping — worth understanding why." if sig == "warn"
                                       else "Promoter stake falling materially — investigate.")})
    elif (shareholding or {}).get("promoter") is not None:
        add(c, "Promoter Holding", "promoter % (NSE)", shareholding["promoter"], fp, None, ("", "", ""))
    else:
        manual(c, "Promoter Holding Trend", "A falling promoter stake warrants investigation. Check NSE shareholding pattern.")
    # -- items genuinely only in annual reports --
    for n, why in [
        ("Related-Party Transactions / Revenue", "High RPTs can hide value leakage. See annual report notes."),
        ("Contingent Liabilities / Net Worth", "Off-balance-sheet obligations. See annual report notes."),
        ("Auditor Qualifications / EoM", "Qualified opinions or emphasis-of-matter are serious. See audit report."),
        ("Subsidiary Loss / Consolidated PAT", "Hidden drag in the group. Compare consolidated vs standalone."),
        ("Promoter Remuneration / PAT", "Excess extraction by management. See governance report."),
    ]:
        manual(c, n, why)

    # ---- tally for the summary chip ----
    counts = {"good": 0, "warn": 0, "bad": 0, "na": 0}
    for cat in cats:
        for r in cat["ratios"]:
            counts[r["signal"]] = counts.get(r["signal"], 0) + 1
    return {"categories": cats, "counts": counts, "wacc": WACC}


def build_personas(info, mos, de, roe, pe, npm, fcf, ocf, ni):
    """Rule-based investor commentary derived from the real metrics."""
    name = info.get("longName", "The company")
    de = de if de is not None else 0
    roe = roe if roe is not None else 0
    pe = pe if pe else 0
    quality = "high" if roe >= 18 else "moderate" if roe >= 10 else "low"
    valuation = "cheap" if mos > 20 else "fair" if mos > -5 else "expensive"
    lev = "debt-heavy" if de > 1 else "moderately geared" if de > 0.3 else "lightly geared"

    def verdict_from(v, q):
        if v == "cheap" and q != "low": return ("Attractive", "green")
        if v == "expensive": return ("Overvalued", "red")
        if q == "high": return ("Quality, watch price", "amber")
        return ("Hold / watch", "amber")

    vb = verdict_from(valuation, quality)

    P = {}
    P["buffett"] = {
        "verdict": vb[0], "vc": vb[1],
        "text": (f"I look for wonderful businesses at fair prices. This one earns a "
                 f"{quality} return on equity and is {lev}. "
                 + ("The balance sheet is the kind of fortress I like to sleep behind. "
                    if de < 0.3 else "The debt load means I'd want to understand the durability of earnings before committing. ")
                 + ("At the current price the market is already paying up for quality — I'd wait for a fearful day. "
                    if valuation == "expensive" else
                    "The price looks reasonable relative to what the business is worth. "
                    if valuation == "fair" else
                    "Price sits below my estimate of value, which is where bargains live. "))
    }
    P["munger"] = {
        "verdict": vb[0], "vc": vb[1],
        "text": ("Invert: what would ruin this? " +
                 ("High leverage in a downturn is the obvious killer here. "
                  if de > 1 else "No single obvious wipe-out risk jumps out on the balance sheet. ")
                 + (f"A {quality} return on capital that needs little reinvestment is a beautiful thing; "
                    if quality == "high" else "Returns on capital are unremarkable, so the business must earn its keep on price. ")
                 + ("but paying a rich multiple for it invites disappointment."
                    if valuation == "expensive" else "and the price isn't asking me to be a hero."))
    }
    graham_v = ("Undervalued", "green") if mos > 33 else ("Borderline", "amber") if mos > 0 else ("Avoid", "red")
    P["graham"] = {
        "verdict": graham_v[0], "vc": graham_v[1],
        "text": (f"My test is margin of safety. Estimated value versus price gives a margin of "
                 f"{mos:+.0f}%. "
                 + ("That clears my one-third cushion — a defensive purchase. "
                    if mos > 33 else "That is a thin cushion, below the one-third I demand; discipline says wait. "
                    if mos > 0 else "The price exceeds appraised value — there is no safety margin, so I pass. ")
                 + (f"A P/E of {pe:.0f} is " + ("within defensive limits." if pe and pe < 20 else "richer than a defensive investor should pay.") if pe else ""))
    }
    lynch_v = ("Buy the dip", "green") if (valuation != "expensive" and quality != "low") else ("Wait", "amber") if valuation == "expensive" else ("Hold", "amber")
    P["lynch"] = {
        "verdict": lynch_v[0], "vc": lynch_v[1],
        "text": ("I want growth at a reasonable price and a story I can explain simply. "
                 + (f"With {quality} returns and a {valuation} price, "
                    + ("this looks like the quiet value I hunt for — accumulate on weakness. "
                       if valuation != "expensive" and quality != "low" else
                       "the story may be good but the price already reflects it — I'd wait for a pullback. "))
                 + ("Cash generation supports the story." if fcf and fcf > 0 else "I'd watch the cash flow — growth that doesn't turn into cash worries me."))
    }
    # Dalio — a SIZING call (not buy/sell): how big a position the risk justifies. Varies by stock.
    _sec = ((info.get("sector") or "") + " " + name).lower()
    fin = any(k in _sec for k in ["bank", "financ", "insurance", "nbfc"])
    cyclical = any(k in _sec for k in ["metal", "steel", "mining", "energy", "oil", "gas",
                                       "cement", "auto", "commodit", "real estate", "realty", "shipping"])
    if de < 0.3 and quality != "low" and not cyclical:
        dalio_v = ("Can carry full weight", "green")
        dalio_act = "sizing OK — a low-debt, steady machine can hold a larger slice; no action needed if already core."
    elif de > 1 or (cyclical and quality == "low"):
        dalio_v = ("Underweight — size down", "amber")
        dalio_act = ("action: keep it SMALL. If it's an oversized slice of your book, trim to a risk-appropriate "
                     "weight (this is a sizing trim, not an exit) — it's rate/cycle-sensitive.")
    else:
        dalio_v = ("Standard weight", "blue")
        dalio_act = "sizing: a normal position — don't oversize it; fine to hold at a middling weight."
    P["dalio"] = {
        "verdict": dalio_v[0], "vc": dalio_v[1],
        "text": ("I think in machines and regimes — my call is how BIG the position should be, not whether to buy or sell. "
                 + (f"As a machine this is a {lev} business with {quality} profitability"
                    + (", in a cyclical/commodity regime" if cyclical else "") + ". ")
                 + ("Low debt means it survives most economic weather. " if de < 0.3
                    else "The leverage makes it rate- and cycle-sensitive. ")
                 + "What this means — " + dalio_act)
    }
    # 6th lens — Institutional Risk / Portfolio Construction (Vanguard/BlackRock/Blackstone tradition).
    # Also a WEIGHT call that varies by the stock's leverage, cyclicality and correlation.
    if de < 0.3 and quality != "low" and not cyclical:
        risk_v = ("Can hold full weight", "green")
        risk_act = "hold — a fortress balance sheet and idiosyncratic drivers let it carry a full weight."
    elif de > 1 or cyclical:
        risk_v = ("Cap the weight — trim if oversized", "amber")
        risk_act = ("action: cap this position. If it already clusters with your other "
                    + ("rate/credit bets" if fin else "cyclical/commodity bets")
                    + " or is an oversized slice, trim it toward a risk-appropriate weight (reduce, don't necessarily exit).")
    else:
        risk_v = ("Moderate weight", "blue")
        risk_act = "hold at a moderate weight; watch that it isn't doubling a factor you already own elsewhere."
    P["risk"] = {
        "verdict": risk_v[0], "vc": risk_v[1],
        "text": ("Position construction, not stock-picking, is my job — this is a WEIGHT call, not a buy/sell. "
                 + (f"A {lev}, {quality}-return business"
                    + (" in a cyclical regime" if cyclical else "") + ". ")
                 + "The real risks are driver-correlation (how much of your book already moves with "
                 + ("rates/credit), " if fin else "rates, capex, commodities, INR, the global cycle), ")
                 + "exit liquidity, and the loss in a severe scenario. What this means — " + risk_act)
    }
    return P


@app.route("/")
def index():
    # serve from this script's own folder, so the app works no matter where the
    # folder lives or what the current working directory is (rename/move-safe).
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), "equity-dashboard.html")


_SEARCH_CACHE = {}   # q -> (timestamp, results)


@app.route("/api/search")
def search():
    """Typeahead across every NSE/BSE-listed company. The curated COMPANIES list
    gives instant hits for the well-known names; Yahoo's search (works from any
    IP) supplies the full universe (e.g. 'Navin' -> Navin Fluorine). One row per
    company (NSE preferred), tagged with the exchange(s) it trades on."""
    import time as _t
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"results": []})
    key = q.lower()
    hit = _SEARCH_CACHE.get(key)
    if hit and (_t.time() - hit[0] < 300):
        return jsonify({"results": hit[1]})

    by_base, order = {}, []

    def put(base, name, exch, query=None):
        if base not in by_base:
            by_base[base] = {"query": query or base, "name": name, "exch": set()}
            order.append(base)
        by_base[base]["exch"].add(exch)
        if exch == "NSE":                       # prefer the NSE ticker for fetching
            by_base[base]["query"] = base
            by_base[base]["name"] = name

    # Yahoo search = the authoritative, always-current NSE/BSE universe (it reflects
    # renames/demergers, e.g. TATAMOTORS -> TMCV/TMPV, so we never suggest a dead
    # ticker). The curated COMPANIES list is used only if Yahoo is unreachable.
    try:
        from curl_cffi import requests as creq
        s = creq.Session(impersonate="chrome")
        r = s.get("https://query1.finance.yahoo.com/v1/finance/search",
                  params={"q": q, "quotesCount": 15, "newsCount": 0, "enableFuzzyQuery": "true"},
                  timeout=8)
        for x in r.json().get("quotes", []):
            symf = str(x.get("symbol", ""))
            if not symf.endswith((".NS", ".BO")) or x.get("quoteType") != "EQUITY":
                continue
            base = symf.rsplit(".", 1)[0]
            exch = "NSE" if symf.endswith(".NS") else "BSE"
            name = (x.get("longname") or x.get("shortname") or base).title()
            put(base, name, exch, query=symf)
    except Exception:
        pass

    if not order:                               # offline / Yahoo down -> curated list
        for name, sym in COMPANIES:
            if key in name.lower() or key in sym.lower():
                put(sym.upper(), name, "NSE")

    results = []
    for base in order:
        e = by_base[base]
        tag = " · ".join([x for x in ("NSE", "BSE") if x in e["exch"]])
        results.append({"symbol": e["query"], "display": base, "name": e["name"], "exch": tag})
    results = results[:8]
    _SEARCH_CACHE[key] = (_t.time(), results)
    return jsonify({"results": results})


@app.route("/api/holdings", methods=["GET", "POST"])
def api_holdings():
    """Persist / restore the Equity Holdings portfolio (symbols, positions, last verdicts)."""
    import json
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "holdings_state.json")
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=1)
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "reason": str(e)}), 200
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return jsonify(json.load(f))
        except Exception:
            pass
    return jsonify({"holdings": []})


@app.route("/api/screener", methods=["POST"])
def api_screener():
    """Parse an uploaded Screener.in Excel export -> quarterly series (₹ Cr)."""
    import screener
    f = request.files.get("file")
    if f is None:
        return jsonify({"error": "No file uploaded."}), 400
    try:
        return jsonify(screener.parse_screener_xlsx(f.read()))
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}", "quarters": []}), 200


@app.route("/api/research")
def api_research():
    """Serve the research/policy overlay (generated by a Claude web-research pass)."""
    import json
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "research_overlay.json")
    if not os.path.exists(path):
        return jsonify({"available": False, "bySymbol": {}})
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        # auto-archive each distinct research version (dated audit trail)
        try:
            asof = str(data.get("asOf", "unknown")).replace(":", "-").split(".")[0][:10]
            adir = os.path.join(os.path.dirname(path), "research_archive")
            os.makedirs(adir, exist_ok=True)
            apath = os.path.join(adir, f"overlay_{asof}.json")
            # keep the latest version for each research date (quarterly cadence)
            with open(apath, "w", encoding="utf-8") as af:
                json.dump(data, af, ensure_ascii=False, indent=1)
        except Exception:
            pass
        data["available"] = True
        return jsonify(data)
    except Exception as e:
        return jsonify({"available": False, "reason": str(e), "bySymbol": {}})


@app.route("/api/framework")
def api_framework():
    # Parsed fresh each call so edits to the Excel reflect on reload.
    try:
        return jsonify(framework.load_framework())
    except Exception as e:
        return jsonify({"available": False, "reason": f"{type(e).__name__}: {e}"}), 200


@app.route("/api/export", methods=["POST"])
def api_export():
    """Build an .xlsx of the consolidated portfolio reading and return it."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    data = request.get_json(silent=True) or {}
    rows = data.get("rows", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio_Reading"
    cols = [("Symbol", "symbol"), ("Sector", "sector"), ("Qty", "qty"), ("Avg Cost", "avg"),
            ("LTP", "ltp"), ("Value", "value"), ("Weight %", "weight"), ("P&L %", "pnlPct"),
            ("Audited Verdict", "verdict"), ("Conviction", "conviction"), ("Basis", "basis"),
            ("Live Status", "liveStatus"), ("Live Signal", "liveAction"),
            ("Live Notes", "liveNotes"), ("Rationale", "rationale")]
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ci, (label, _) in enumerate(cols, 1):
        c = ws.cell(1, ci, label); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, (_, key) in enumerate(cols, 1):
            v = row.get(key)
            if key in ("weight", "pnlPct") and isinstance(v, (int, float)):
                v = round(v, 1)
            if key in ("value", "avg", "ltp") and isinstance(v, (int, float)):
                v = round(v)
            ws.cell(ri, ci, v)
    widths = [13, 16, 7, 10, 10, 12, 9, 8, 18, 11, 7, 12, 16, 46, 60]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}1"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Portfolio_Reading.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _overlay_for(sym):
    """Return the research overlay entry for a symbol (base ticker), or None."""
    import json
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "research_overlay.json")
        if not os.path.exists(path):
            return None
        with open(path, encoding="utf-8") as f:
            by = json.load(f).get("bySymbol", {})
        base = sym.upper().split(".")[0]
        return by.get(base)
    except Exception:
        return None


def _json_safe(obj):
    """Recursively replace NaN/Infinity with None — those are not valid JSON and
    a single one anywhere makes the whole response unparseable in the browser."""
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else None
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    return obj


@app.route("/api/company/<sym>")
def company(sym):
    try:
        data = build_company(sym)
        if data is None:
            return jsonify({"error": f"Could not fetch data for '{sym}'. Try the NSE symbol (e.g. RELIANCE) or add .BO for BSE."}), 404
        rv = _overlay_for(sym)
        if rv:
            data["research"] = rv
        return jsonify(_json_safe(data))
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500


if __name__ == "__main__":
    print("Equity Dashboard  ->  http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
