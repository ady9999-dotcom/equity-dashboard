"""
Parse a Screener.in "Export to Excel" workbook -> quarterly series.

Screener's export has a "Data Sheet" with a stacked layout; the "Quarters"
section lists ~12 CONSECUTIVE quarters (no gaps) with rows for Sales,
Operating Profit (= EBITDA), Interest and Net Profit. Values are already in
₹ Crore, so no unit conversion is needed. This is used only to fill quarters
that Yahoo/NSE leave missing on the single-company dashboard.
"""
import io
import re
import datetime

_MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
           "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def _to_ym(v):
    """A cell value -> (year, month) if it looks like a period, else None."""
    if isinstance(v, datetime.datetime):
        return (v.year, v.month)
    if isinstance(v, str):
        s = v.strip()
        m = re.search(r"([A-Za-z]{3})[^0-9]*(\d{2,4})", s)
        if m:
            mo = _MONTHS.get(m.group(1).lower())
            yr = int(m.group(2))
            if yr < 100:
                yr += 2000
            if mo:
                return (yr, mo)
        m = re.search(r"(\d{4})[-/](\d{1,2})", s)
        if m:
            return (int(m.group(1)), int(m.group(2)))
    return None


def _num(v):
    try:
        if v in (None, ""):
            return None
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def parse_screener_xlsx(data):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["Data Sheet"] if "Data Sheet" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    # company name: Screener puts label "COMPANY NAME" in A1 and the name in B1
    company = None
    for r in rows[:6]:
        if r and r[0] and str(r[0]).strip().upper() == "COMPANY NAME":
            company = str(r[1]).strip() if (len(r) > 1 and r[1]) else None
            break
    if not company:
        company = next((str(r[0]).strip() for r in rows[:3] if r and r[0]), None)

    def c0(r):
        return str(r[0]).strip().lower() if (r and r[0] is not None) else ""

    # locate the "Quarters" section marker
    qstart = next((i for i, r in enumerate(rows) if c0(r) == "quarters"), None)
    if qstart is None:
        return {"company": company, "quarters": [], "error": "No 'Quarters' section found in the Data Sheet."}

    # the quarter date header row (>= 3 date-parseable cells) just below it
    dater = None
    datecols = []
    for i in range(qstart, min(qstart + 6, len(rows))):
        parsed = [(_to_ym(v), j + 1) for j, v in enumerate(rows[i][1:])]
        hits = [(d, col) for d, col in parsed if d]
        if len(hits) >= 3:
            dater = i
            datecols = hits
            break
    if dater is None:
        return {"company": company, "quarters": [], "error": "Could not find the quarterly date row."}

    # section ends at the next block header
    end = len(rows)
    for i in range(dater + 1, len(rows)):
        up = str(rows[i][0]).strip().upper() if (rows[i] and rows[i][0]) else ""
        if i > dater + 3 and up in ("BALANCE SHEET", "CASH FLOW", "PROFIT & LOSS", "DERIVED", "RATIOS"):
            end = i
            break

    def findrow(*prefixes):
        for i in range(dater + 1, end):
            label = c0(rows[i])
            if any(label == p or label.startswith(p) for p in prefixes):
                return rows[i]
        return None

    sales = findrow("sales", "revenue")
    op = findrow("operating profit")
    interest = findrow("interest")
    net = findrow("net profit")

    def val(r, col):
        return _num(r[col]) if (r and col < len(r)) else None

    quarters = []
    for (yr, mo), col in datecols:
        quarters.append({"y": yr, "m": mo,
                         "rev": val(sales, col), "ebitda": val(op, col),
                         "interest": val(interest, col), "np": val(net, col)})
    return {"company": company, "quarters": quarters}
