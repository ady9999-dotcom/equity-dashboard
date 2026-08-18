"""
Reads the investor's audited portfolio framework (the multi-tab Excel produced by
the Master Project Prompt) into JSON keyed by symbol, so the dashboard's
"Equity Holdings" tab can merge it with an uploaded Zerodha holdings CSV.

The Excel is the *audited output* of the framework (verdict, conviction, basis,
rationale per holding) enriched with sector-policy, risk-register, valuation and
tax (loss set-off) context. This module does not re-derive verdicts — it surfaces
the analyst's, exactly as filed, with their as-of dates and provisional flags.
"""
import os
import re
import glob
import shutil
import tempfile
import datetime as dt

try:
    import openpyxl
    _HAVE = True
except Exception:
    _HAVE = False

def _framework_folder():
    """Folder holding your audited-portfolio Excel. Resolved without hard-coding any
    personal path in the source (so this is safe to publish):
      1. env var EQUITY_FRAMEWORK_XLSX points straight at the .xlsx (handled below)
      2. a local, git-ignored 'framework_path.txt' next to this file, containing the folder
      3. a './portfolio' folder next to this file
    Put your own folder path in framework_path.txt (it is never committed)."""
    here = os.path.dirname(os.path.abspath(__file__))
    cfg = os.path.join(here, "framework_path.txt")
    if os.path.exists(cfg):
        try:
            p = open(cfg, encoding="utf-8").read().strip()
            if p:
                return p
        except Exception:
            pass
    return os.path.join(here, "portfolio")


def _locate_latest():
    """Find the newest consolidated master (append-only dated files, e.g.
    SU3550_Consolidated_Master_YYYYMMDD.xlsx); fall back to the older
    'Audited Portfolio.xlsx'. Override with EQUITY_FRAMEWORK_XLSX."""
    env = os.environ.get("EQUITY_FRAMEWORK_XLSX")
    if env:
        return env
    folder = _framework_folder()
    cands = [c for c in glob.glob(os.path.join(folder, "*Consolidated_Master*.xlsx"))
             if not os.path.basename(c).startswith("~$")]
    if cands:
        return max(cands, key=os.path.getmtime)
    return os.path.join(folder, "Audited Portfolio.xlsx")


# Kept for backward-compatibility (some callers import DEFAULT_PATH).
DEFAULT_PATH = _locate_latest()


def _s(v):
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.strftime("%d-%b-%Y")
    return str(v).strip()


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def load_framework(path=None):
    """Parse the workbook -> dict, or {'available': False, ...} if unreadable.
    Re-locates the newest master each call so a fresh dated file is picked up
    automatically; tolerates the file being open in Excel (copies it first)."""
    path = path or _locate_latest()
    if not _HAVE:
        return {"available": False, "reason": "openpyxl not installed on the server."}
    if not os.path.exists(path):
        return {"available": False, "reason": f"Framework Excel not found at: {path}",
                "path": path}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except PermissionError:
        # file is open in Excel — read a temp copy instead
        try:
            tmp = os.path.join(tempfile.gettempdir(), "framework_snapshot.xlsx")
            shutil.copy2(path, tmp)
            wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)
        except Exception as e:
            return {"available": False, "reason": f"Could not open Excel (in use): {e}", "path": path}
    except Exception as e:
        return {"available": False, "reason": f"Could not open Excel: {e}", "path": path}

    sheets = wb.sheetnames
    by = {}          # symbol -> record
    order = []

    # ---- Master_Holdings: identity ----
    if "Master_Holdings" in sheets:
        for r in _rows(wb["Master_Holdings"])[1:]:
            sym = _s(r[0]).upper()
            if not sym:
                continue
            by.setdefault(sym, {"symbol": sym})
            by[sym].update({
                "sector": _s(r[1]) if len(r) > 1 else "",
                "xlQty": _f(r[2]) if len(r) > 2 else None,
                "xlQtyLong": _f(r[3]) if len(r) > 3 else None,
                "xlAvgCost": _f(r[4]) if len(r) > 4 else None,
                "refPrice": _f(r[5]) if len(r) > 5 else None,
                "priceAsOf": _s(r[6]) if len(r) > 6 else "",
            })
            if sym not in order:
                order.append(sym)

    ALLSYMS = set(by.keys())

    def match_syms(text):
        t = _s(text)
        return [s for s in ALLSYMS if re.search(r"(?<![A-Z0-9])" + re.escape(s) + r"(?![A-Z0-9])", t)]

    # ---- Portfolio_Action_ALL (NEW master schema): the core call + position ----
    # cols: Symbol, Qty, Avg Cost, LTP, Cur Value, Unreal P&L, P&L %, Verdict, Conv,
    #       Depth, LT-qualify, Action Now, Buy/Add Band, Rationale (current-dated)
    if "Portfolio_Action_ALL" in sheets:
        for r in _rows(wb["Portfolio_Action_ALL"])[1:]:
            sym = _s(r[0]).upper()
            if not sym or sym == "SYMBOL":
                continue
            rec = by.setdefault(sym, {"symbol": sym})
            if sym not in order:
                order.append(sym)
            rec["xlQty"] = _f(r[1]) if len(r) > 1 else rec.get("xlQty")
            rec["xlAvgCost"] = _f(r[2]) if len(r) > 2 else rec.get("xlAvgCost")
            rec["refPrice"] = _f(r[3]) if len(r) > 3 else rec.get("refPrice")
            rec["verdict"] = _s(r[7]) if len(r) > 7 else ""
            rec["conviction"] = _f(r[8]) if len(r) > 8 else None
            rec["basis"] = _s(r[9]) if len(r) > 9 else ""
            rec["ltQualify"] = _s(r[10]) if len(r) > 10 else ""
            rec["actionNow"] = _s(r[11]) if len(r) > 11 else ""
            rec["buyBand"] = _s(r[12]) if len(r) > 12 else ""
            rec["rationale"] = _s(r[13]) if len(r) > 13 else ""

    # ---- Verdict_Action (OLD schema fallback) ----
    if "Verdict_Action" in sheets:
        for r in _rows(wb["Verdict_Action"])[1:]:
            sym = _s(r[0]).upper()
            if not sym:
                continue
            rec = by.setdefault(sym, {"symbol": sym})
            if sym not in order:
                order.append(sym)
            rec.setdefault("verdict", _s(r[1]) if len(r) > 1 else "")
            rec.setdefault("conviction", _f(r[2]) if len(r) > 2 else None)
            rec.setdefault("basis", _s(r[3]) if len(r) > 3 else "")
            rec.setdefault("rationale", _s(r[4]) if len(r) > 4 else "")

    # ---- Valuation_ALL (NEW): the MoS Buy Band + in-band + buy note the user asked for ----
    # cols: Symbol, Depth, MoS Buy Band, LTP, In band?, Conv, Buy note
    if "Valuation_ALL" in sheets:
        for r in _rows(wb["Valuation_ALL"])[1:]:
            sym = _s(r[0]).upper()
            if sym in by and sym != "SYMBOL":
                by[sym]["valuation"] = {
                    "depth": _s(r[1]) if len(r) > 1 else "",
                    "mosBand": _s(r[2]) if len(r) > 2 else "",
                    "refPrice": _s(r[3]) if len(r) > 3 else "",
                    "inBand": _s(r[4]) if len(r) > 4 else "",
                    "conv": _s(r[5]) if len(r) > 5 else "",
                    "buyNote": _s(r[6]) if len(r) > 6 else "",
                }

    # ---- Valuation (OLD schema fallback: bear/base/bull IV) ----
    if "Valuation" in sheets:
        for r in _rows(wb["Valuation"])[1:]:
            sym = _s(r[0]).upper()
            if sym in by and "valuation" not in by[sym]:
                by[sym]["valuation"] = {
                    "method": _s(r[1]), "bearIV": _f(r[2]), "baseIV": _f(r[3]),
                    "bullIV": _f(r[4]), "mosBand": _s(r[5]), "refPrice": _s(r[6]),
                    "vsBase": _s(r[7]) if len(r) > 7 else "",
                    "reverseDCF": _s(r[8]) if len(r) > 8 else "",
                    "asOf": _s(r[9]) if len(r) > 9 else "",
                }

    # ---- Loss_SetOff_Register: tax note ----
    if "Loss_SetOff_Register" in sheets:
        for r in _rows(wb["Loss_SetOff_Register"])[1:]:
            sym = _s(r[0]).upper()
            if sym in by:
                by[sym]["tax"] = {
                    "loss": _f(r[1]), "period": _s(r[2]),
                    "action": _s(r[3]) if len(r) > 3 else "",
                    "note": _s(r[4]) if len(r) > 4 else "",
                }

    # ---- Exit_Trim_Schedule: execution plan ----
    if "Exit_Trim_Schedule" in sheets:
        for r in _rows(wb["Exit_Trim_Schedule"])[1:]:
            sym = _s(r[1]).upper()
            if sym in by:
                by[sym].setdefault("exit", []).append({
                    "priority": _s(r[0]), "action": _s(r[2]),
                    "status": _s(r[3]) if len(r) > 3 else "",
                    "qty": _s(r[4]) if len(r) > 4 else "",
                    "priceLevel": _s(r[7]) if len(r) > 7 else "",
                    "estPnl": _s(r[9]) if len(r) > 9 else "",
                    "notes": _s(r[11]) if len(r) > 11 else "",
                })

    # ---- Quarterly_Tracking: latest thesis status ----
    if "Quarterly_Tracking" in sheets:
        for r in _rows(wb["Quarterly_Tracking"])[1:]:
            sym = _s(r[1]).upper()
            if sym in by:
                by[sym].setdefault("quarterly", []).append({
                    "quarter": _s(r[0]), "revDelta": _s(r[2]), "patDelta": _s(r[3]),
                    "metric": _s(r[4]), "guidance": _s(r[5]) if len(r) > 5 else "",
                    "thesis": _s(r[6]) if len(r) > 6 else "",
                    "verdict": _s(r[7]) if len(r) > 7 else "",
                    "action": _s(r[8]) if len(r) > 8 else "",
                })

    # ---- Sector_Policy_Map: tag each affected holding ----
    policies = []
    if "Sector_Policy_Map" in sheets:
        for r in _rows(wb["Sector_Policy_Map"])[1:]:
            fact, tag, affected, logic = (_s(r[0]), _s(r[1]),
                                          _s(r[2]) if len(r) > 2 else "",
                                          _s(r[3]) if len(r) > 3 else "")
            if not fact:
                continue
            hits = match_syms(affected)
            policies.append({"fact": fact, "tag": tag, "affected": hits, "logic": logic})
            for s in hits:
                by[s].setdefault("policies", []).append({"fact": fact, "tag": tag, "logic": logic})

    # ---- Risk_Register: attach risks to holdings ----
    risks = []
    if "Risk_Register" in sheets:
        for r in _rows(wb["Risk_Register"])[1:]:
            risk = _s(r[1]) if len(r) > 1 else ""
            if not risk:
                continue
            affected = _s(r[2]) if len(r) > 2 else ""
            hits = match_syms(affected)
            item = {"risk": risk, "affectedText": affected, "affected": hits,
                    "type": _s(r[3]) if len(r) > 3 else "",
                    "L": _s(r[4]) if len(r) > 4 else "", "I": _s(r[5]) if len(r) > 5 else "",
                    "mitigation": _s(r[6]) if len(r) > 6 else ""}
            risks.append(item)
            for s in hits:
                by[s].setdefault("risks", []).append(item)

    # ---- README meta (new master: "README"; old: "README_Audit") ----
    meta = {"readme": []}
    _readme_sheet = "README_Audit" if "README_Audit" in sheets else ("README" if "README" in sheets else None)
    if _readme_sheet:
        for r in _rows(wb[_readme_sheet])[1:]:
            if len(r) >= 3 and _s(r[1]):
                meta["readme"].append({"item": _s(r[1]), "detail": _s(r[2])})
                low = _s(r[1]).lower()
                if "as-of" in low or "data as" in low:
                    meta["asOf"] = _s(r[2])
                if "basis legend" in low:
                    meta["basisLegend"] = _s(r[2])
                if "disclaimer" in low:
                    meta["disclaimer"] = _s(r[2])

    return {
        "available": True,
        "path": path,
        "meta": meta,
        "bySymbol": by,
        "order": order,
        "policies": policies,
        "risks": risks,
        "count": len(by),
    }


if __name__ == "__main__":
    fw = load_framework()
    if not fw.get("available"):
        print("NOT AVAILABLE:", fw.get("reason"))
    else:
        print("holdings:", fw["count"], "| policies:", len(fw["policies"]), "| risks:", len(fw["risks"]))
        for s in ["BALRAMCHIN", "HIKAL", "RELIANCE"]:
            r = fw["bySymbol"].get(s, {})
            print(f"\n{s}: {r.get('verdict')} conv={r.get('conviction')} basis={r.get('basis')}")
            print("  rationale:", (r.get("rationale") or "")[:90])
            print("  policies:", len(r.get("policies", [])), "risks:", len(r.get("risks", [])),
                  "valuation:", bool(r.get("valuation")), "tax:", bool(r.get("tax")), "exit:", len(r.get("exit", [])))
